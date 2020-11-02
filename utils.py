# -*- coding: utf-8 -*-
"""
Created on Thu Mar  5 10:56:11 2020

@author: Getachew
"""

#this script was written to automate data entry of ethio-chicken data from excel files

from openpyxl import load_workbook
#openpyxl.__version__
# 3.0.0

import pandas as pd
pd.__version__
#1.0.1

import numpy as np
np.__version__
#'1.16.5'

import pandas as pd
import re


#####################################################3
def get_data1(file_path, files):
    [
    pa_id,
    name_assessor,
    date,
    start_time,  
    end_time,
    duration,
    type_assessed,
    country,
    code,
    business_name,
    legal_formation,
    region,
    zone,
    woreda,
    telephone,
    email,
    contact_person,
    contact_designation,
    contact_gender,
    contact_educ,
    hh_head_owner,
    train_given_by,
    train_satisfaction,
    train_decision,
    have_license,
    have_tin,
    year_estabilished,
    employment_male, 
    employment_female, 
    employment_total,
    avail_water, 
    avail_input_store, 
    avail_electricity, 
    avail_generator, 
    avail_heating, 
    ss_water_man, 
    ss_water_woman, 
    ss_water_children, 
    ss_water_all,
    source_prot_inputs,
    get_business_advice, 
    access_poultry_mart, 
    agency_access_loan, 
    agency_lender_type,
    agency_amount, 
    informal_access_loan, 
    informal_lender_type, 
    informal_amount,
    finance_challenge1, 
    finance_challenge2, 
    finance_challenge3,
    finance_challenge4,
    finance_challenge_other, 
    
    level_educ, 
    experience_years, 
    received_training, 
    employment_procedure, 
    provide_advisory_serv, 
    convenient_poultry_house,
    know_feed_types,
    source_of_water, 
    quality_attribute, 
    poultry_disease,
    know_poultry_prtection, 
    know_hygen, 
    have_annual_plan, 
    have_additional_income,  
    alternative_finance, 
    market_competition, 
    promotion, 
    risk_price, 
    risk_ss_shortage,
    working_capital, 
    financial_records, 
    risk_potential,
    mitigation_method, 
    risk_health,
    
    prod_cycle2016, 
    prod_cycle2017, 
    prod_cycle2018,
    num_chicks_sold_farm2018, 
    num_chicks_sold_gov_agent2018,
    num_chicks_sold_trader2018, 
    num_chicks_sold_vill_agents2018,
    num_chicks_sold_other2018, 
    num_chicks_sold_total2018,
    num_chicks_sold_farm2017, 
    num_chicks_sold_gov_agent2017,
    num_chicks_sold_trader2017, 
    num_chicks_sold_vill_agents2017,
    num_chicks_sold_other2017, 
    num_chicks_sold_total2017,
    num_chicks_sold_farm2016, 
    num_chicks_sold_gov_agent2016,
    num_chicks_sold_trader2016, 
    num_chicks_sold_vill_agents2016,
    num_chicks_sold_other2016, 
    num_chicks_sold_total2016,
    turn_over_farm2018, 
    turn_over_gov_agent2018,
    turn_over_trader2018, 
    turn_over_vill_agents2018,
    turn_over_other2018, 
    turn_over_total2018,
    turn_over_farm2017, 
    turn_over_gov_agent2017,
    turn_over_trader2017, 
    turn_over_vill_agents2017,
    turn_over_other2017, 
    turn_over_total2017,
    turn_over_farm2016, 
    turn_over_gov_agent2016,
    turn_over_trader2016,
    turn_over_vill_agents2016,
    turn_over_other2016,
    turn_over_total2016, 
    sold_to_unique_farmers2018,
    sold_to_unique_farmers2017,
    sold_to_unique_farmers2016,
    
    mgmt_capability_lev,
    inst_competency_lev,
    bus_knowledge_lev,
    mark_potential_lev,
    fin_management_lev,
    env_compliance_lev,
    mgmt_capability_per,
    inst_competency_per,
    bus_knowledge_per,
    mark_potential_per,
    fin_management_per,
    env_compliance_per
    ]= ([] for i in range(131))
    
    for file in files:
        match = re.search(r'\d*', file)
        pa_id.append(match.group())
        
        workbook = load_workbook(file_path+file, data_only=True)
                    
        worksheet = workbook['BIETH']
        print(f'Working on file: {file} ...')
        print('Sheet: BIETH ...')
    
        name_assessor_obj = worksheet['G7']
        date_obj  = worksheet['A11']
        start_time_obj  = worksheet['B11']
        end_time_obj = worksheet['C11']
        duration_obj = worksheet['D11']
        type_assessed_obj = worksheet['F10']
        country_obj = worksheet['G10']
        code_obj = worksheet['H10']
        business_name_obj = worksheet['C15']
        legal_formation_obj = worksheet['C16']
        region_obj = worksheet['C17']
        zone_obj = worksheet['C18']
        woreda_obj = worksheet['C19']
        telephone_obj = worksheet['C20']
        email_obj = worksheet['C21']
        contact_person_obj = worksheet['G15']
        contact_designation_obj = worksheet['H15']
        contact_gender_obj = worksheet['G16']
        contact_educ_obj = worksheet['G17']
        hh_head_owner_obj = worksheet['G19']
        train_given_by_obj = worksheet['C24']
        train_satisfaction_obj = worksheet['F24']
        train_decision_obj = worksheet['G24']
        have_license_obj = worksheet['C30']
        have_tin_obj = worksheet['C31']
        year_estabilished_obj = worksheet['G30']
        employment_male_obj = worksheet['A36']
        employment_female_obj = worksheet['B36']
        employment_total_obj = worksheet['C36']
        avail_water_obj = worksheet['H33']
        avail_input_store_obj = worksheet['H34']
        avail_electricity_obj = worksheet['H35']
        avail_generator_obj = worksheet['H36']
        avail_heating_obj = worksheet['H37']
        ss_water_man_obj = worksheet['C40']
        ss_water_woman_obj = worksheet['F40']
        ss_water_children_obj = worksheet['G40']
        ss_water_all_obj = worksheet['H40']
        source_prot_inputs_obj=worksheet['A43']
        get_business_advice_obj = worksheet['C43']
        access_poultry_mart_obj = worksheet['F43']
        agency_access_loan_obj = worksheet['C45']
        agency_lender_type_obj = worksheet['F45']
        agency_amount_obj = worksheet['H45']
        informal_access_loan_obj = worksheet['C46']
        informal_lender_type_obj = worksheet['F46'] 
        informal_amount_obj = worksheet['H46']
        finance_challenge1_obj = worksheet['A49']
        finance_challenge2_obj = worksheet['C49']
        finance_challenge3_obj = worksheet['E49']
        finance_challenge4_obj = worksheet['G49']
        finance_challenge_other_obj = worksheet['D50']
    
        worksheet = workbook['QETH']
        print(f'Working on file: {file} ...')
        print('Sheet: QETH ...')
    
        level_educ_obj = worksheet['F14']
        experience_years_obj = worksheet['F20']
        received_training_obj = worksheet['F26']
        employment_procedure_obj = worksheet['F33']
        provide_advisory_serv_obj = worksheet['F39']
        convenient_poultry_house_obj = worksheet['F49']
        know_feed_types_obj = worksheet['F55']    
        source_of_water_obj = worksheet['F61']
        quality_attribute_obj = worksheet['F67']
        poultry_disease_obj = worksheet['F73']
        know_poultry_prtection_obj = worksheet['F79']
        know_hygen_obj = worksheet['F86']
        have_annual_plan_obj = worksheet['F96']
        have_additional_income_obj = worksheet['F103']
        alternative_finance_obj = worksheet['F110']
        market_competition_obj = worksheet['F120']
        promotion_obj = worksheet['F126']
        risk_price_obj = worksheet['F133']
        risk_ss_shortage_obj = worksheet['F139']
        working_capital_obj = worksheet['F149']
        financial_records_obj = worksheet['F155']
        risk_potential_obj = worksheet['F165']
        mitigation_method_obj = worksheet['F171']
        risk_health_obj = worksheet['F178']
    
        worksheet = workbook['Additional Info']
        print(f'Working on file: {file} ...')
        print('Sheet: Additional Info ...')
    
        prod_cycle2016_obj = worksheet['C14']
        prod_cycle2017_obj = worksheet['D14']
        prod_cycle2018_obj = worksheet['E14']
        
        num_chicks_sold_farm2018_obj = worksheet['C19']
        num_chicks_sold_gov_agent2018_obj = worksheet['D19']
        num_chicks_sold_trader2018_obj = worksheet['E19']
        num_chicks_sold_vill_agents2018_obj = worksheet['F19']
        num_chicks_sold_other2018_obj = worksheet['G19']
        num_chicks_sold_total2018_obj = worksheet['H19']
        
        turn_over_farm2018_obj = worksheet['C20']
        turn_over_gov_agent2018_obj = worksheet['D20']
        turn_over_trader2018_obj = worksheet['E20']
        turn_over_vill_agents2018_obj = worksheet['F20']
        turn_over_other2018_obj = worksheet['G20']
        turn_over_total2018_obj = worksheet['H20']
        
        sold_to_unique_farmers2018_obj = worksheet['G21']
    
        num_chicks_sold_farm2017_obj = worksheet['C25']
        num_chicks_sold_gov_agent2017_obj = worksheet['D25']
        num_chicks_sold_trader2017_obj = worksheet['E25']
        num_chicks_sold_vill_agents2017_obj = worksheet['F25']
        num_chicks_sold_other2017_obj = worksheet['G25']
        num_chicks_sold_total2017_obj = worksheet['H25']
        
        turn_over_farm2017_obj = worksheet['C26']
        turn_over_gov_agent2017_obj = worksheet['D26']
        turn_over_trader2017_obj = worksheet['E26']
        turn_over_vill_agents2017_obj = worksheet['F26']
        turn_over_other2017_obj = worksheet['G26']
        turn_over_total2017_obj = worksheet['H26']
        
        sold_to_unique_farmers2017_obj = worksheet['G27']
    
        
        num_chicks_sold_farm2016_obj = worksheet['C31']
        num_chicks_sold_gov_agent2016_obj = worksheet['D31']
        num_chicks_sold_trader2016_obj = worksheet['E31']
        num_chicks_sold_vill_agents2016_obj = worksheet['F31']
        num_chicks_sold_other2016_obj = worksheet['G31']
        num_chicks_sold_total2016_obj = worksheet['H3']
        
        
        turn_over_farm2016_obj = worksheet['C32']
        turn_over_gov_agent2016_obj = worksheet['D32']
        turn_over_trader2016_obj = worksheet['E32']
        turn_over_vill_agents2016_obj = worksheet['F32']
        turn_over_other2016_obj = worksheet['G32']
        turn_over_total2016_obj =  worksheet['H32']
    
        sold_to_unique_farmers2016_obj = worksheet['G33']
             
        
        worksheet = workbook['RETH-EN']
        print(f'Working on file: {file} ...')
        print('Sheet: RETH-EN ...')
        
        mgmt_capability_lev_obj=worksheet['D15']
        inst_competency_lev_obj=worksheet['D16']
        bus_knowledge_lev_obj=worksheet['D17']
        mark_potential_lev_obj=worksheet['D18']
        fin_management_lev_obj=worksheet['D19']
        env_compliance_lev_obj=worksheet['D20']
        mgmt_capability_per_obj=worksheet['E15']
        inst_competency_per_obj=worksheet['E16']
        bus_knowledge_per_obj=worksheet['E17']
        mark_potential_per_obj=worksheet['E18']
        fin_management_per_obj=worksheet['E19']
        env_compliance_per_obj=worksheet['E20']
        
        name_assessor.append(name_assessor_obj.value)
        date.append(date_obj.value)
        start_time.append(start_time_obj.value)
        end_time.append(end_time_obj.value)
        duration.append(duration_obj.value)
        type_assessed.append(type_assessed_obj.value)
        country.append(country_obj.value)
        code.append(code_obj.value)
        business_name.append(business_name_obj.value)
        legal_formation.append(legal_formation_obj.value)
        region.append(region_obj.value)
        zone.append(zone_obj.value)
        woreda.append(woreda_obj.value)
        telephone.append(telephone_obj.value)
        email.append(email_obj.value)
        contact_person.append(contact_person_obj.value)
        contact_designation.append(contact_designation_obj.value)
        contact_gender.append(contact_gender_obj.value)
        contact_educ.append(contact_educ_obj.value)
        hh_head_owner.append(hh_head_owner_obj.value)
        train_given_by.append(train_given_by_obj.value)
        train_satisfaction.append(train_satisfaction_obj.value)
        train_decision.append(train_decision_obj.value)
        have_license.append(have_license_obj.value)
        have_tin.append(have_tin_obj.value)
        year_estabilished.append(year_estabilished_obj.value)
        employment_male.append(employment_male_obj.value)
        employment_female.append(employment_female_obj.value)
        employment_total.append(employment_total_obj.value)
        avail_water.append(avail_water_obj.value)
        avail_input_store.append(avail_input_store_obj.value)
        avail_electricity.append(avail_electricity_obj.value)
        avail_generator.append(avail_generator_obj.value)
        avail_heating.append(avail_heating_obj.value)
        ss_water_man.append(ss_water_man_obj.value)
        ss_water_woman.append(ss_water_woman_obj.value)
        ss_water_children.append(ss_water_children_obj.value)
        ss_water_all.append(ss_water_all_obj.value)
        source_prot_inputs.append(source_prot_inputs_obj.value)
        get_business_advice.append(get_business_advice_obj.value)
        access_poultry_mart.append(access_poultry_mart_obj.value)
        agency_access_loan.append(agency_access_loan_obj.value)
        agency_lender_type.append(agency_lender_type_obj.value)
        agency_amount.append(agency_amount_obj.value)
        informal_access_loan.append(informal_access_loan_obj.value)
        informal_lender_type.append(informal_lender_type_obj.value)
        informal_amount.append(informal_amount_obj.value)
        finance_challenge1.append(finance_challenge1_obj.value)
        finance_challenge2.append(finance_challenge2_obj.value)
        finance_challenge3.append(finance_challenge3_obj.value)
        finance_challenge4.append(finance_challenge4_obj.value)
        finance_challenge_other.append(finance_challenge_other_obj.value)
        
        level_educ.append(level_educ_obj.value)
        experience_years.append(experience_years_obj.value)
        received_training.append(received_training_obj.value)
        employment_procedure.append(employment_procedure_obj.value)
        provide_advisory_serv.append(provide_advisory_serv_obj.value)
        convenient_poultry_house.append(convenient_poultry_house_obj.value)
        know_feed_types.append(know_feed_types_obj.value)
        source_of_water.append(source_of_water_obj.value)
        quality_attribute.append(quality_attribute_obj.value)
        poultry_disease.append(poultry_disease_obj.value)
        know_poultry_prtection.append(know_poultry_prtection_obj.value)
        know_hygen.append(know_hygen_obj.value)
        have_annual_plan.append(have_annual_plan_obj.value)
        have_additional_income.append(have_additional_income_obj.value)
        alternative_finance.append(alternative_finance_obj.value)
        market_competition.append(market_competition_obj.value)
        promotion.append(promotion_obj.value)
        risk_price.append(risk_price_obj.value)
        risk_ss_shortage.append(risk_ss_shortage_obj.value)
        working_capital.append(working_capital_obj.value)
        financial_records.append(financial_records_obj.value)
        risk_potential.append(risk_potential_obj.value)
        mitigation_method.append(mitigation_method_obj.value)
        risk_health.append(risk_health_obj.value)
        
        prod_cycle2016.append(prod_cycle2016_obj.value)
        prod_cycle2017.append(prod_cycle2017_obj.value)
        prod_cycle2018.append(prod_cycle2018_obj.value)
        num_chicks_sold_farm2018.append(num_chicks_sold_farm2018_obj.value)
        num_chicks_sold_gov_agent2018.append(num_chicks_sold_gov_agent2018_obj.value)
        num_chicks_sold_trader2018.append(num_chicks_sold_trader2018_obj.value)
        num_chicks_sold_vill_agents2018.append(num_chicks_sold_vill_agents2018_obj.value)
        num_chicks_sold_other2018.append(num_chicks_sold_other2018_obj.value)
        num_chicks_sold_total2018.append(num_chicks_sold_total2018_obj.value)
        num_chicks_sold_farm2017.append(num_chicks_sold_farm2017_obj.value)
        num_chicks_sold_gov_agent2017.append(num_chicks_sold_gov_agent2017_obj.value)
        num_chicks_sold_trader2017.append(num_chicks_sold_trader2017_obj.value)
        num_chicks_sold_vill_agents2017.append(num_chicks_sold_vill_agents2017_obj.value)
        num_chicks_sold_other2017.append(num_chicks_sold_other2017_obj.value)
        num_chicks_sold_total2017.append(num_chicks_sold_total2017_obj.value)
        num_chicks_sold_farm2016.append(num_chicks_sold_farm2016_obj.value)
        num_chicks_sold_gov_agent2016.append(num_chicks_sold_gov_agent2016_obj.value)
        num_chicks_sold_trader2016.append(num_chicks_sold_trader2016_obj.value)
        num_chicks_sold_vill_agents2016.append(num_chicks_sold_vill_agents2016_obj.value)
        num_chicks_sold_other2016.append(num_chicks_sold_other2016_obj.value)
        num_chicks_sold_total2016.append(num_chicks_sold_total2016_obj.value)
        turn_over_farm2018.append(turn_over_farm2018_obj.value)
        turn_over_gov_agent2018.append(turn_over_gov_agent2018_obj.value)
        turn_over_trader2018.append(turn_over_trader2018_obj.value)
        turn_over_vill_agents2018.append(turn_over_vill_agents2018_obj.value)
        turn_over_other2018.append(turn_over_other2018_obj.value)
        turn_over_total2018.append(turn_over_total2018_obj.value)
        turn_over_farm2017.append(turn_over_farm2017_obj.value)
        turn_over_gov_agent2017.append(turn_over_gov_agent2017_obj.value)
        turn_over_trader2017.append(turn_over_trader2017_obj.value)
        turn_over_vill_agents2017.append(turn_over_vill_agents2017_obj.value)
        turn_over_other2017.append(turn_over_other2017_obj.value)
        turn_over_total2017.append(turn_over_total2017_obj.value)
        turn_over_farm2016.append(turn_over_farm2016_obj.value)
        turn_over_gov_agent2016.append(turn_over_gov_agent2016_obj.value)
        turn_over_trader2016.append(turn_over_trader2016_obj.value)
        turn_over_vill_agents2016.append(turn_over_vill_agents2016_obj.value)
        turn_over_other2016.append(turn_over_other2016_obj.value)
        turn_over_total2016.append(turn_over_total2016_obj.value)
        sold_to_unique_farmers2018.append(sold_to_unique_farmers2018_obj.value)
        sold_to_unique_farmers2017.append(sold_to_unique_farmers2017_obj.value)
        sold_to_unique_farmers2016.append(sold_to_unique_farmers2016_obj.value)
    
        mgmt_capability_lev.append(mgmt_capability_lev_obj.value)
        inst_competency_lev.append(inst_competency_lev_obj.value)
        bus_knowledge_lev.append(bus_knowledge_lev_obj.value)
        mark_potential_lev.append(mark_potential_lev_obj.value)
        fin_management_lev.append(fin_management_lev_obj.value)
        env_compliance_lev.append(env_compliance_lev_obj.value)
        mgmt_capability_per.append(mgmt_capability_per_obj.value)
        inst_competency_per.append(inst_competency_per_obj.value)
        bus_knowledge_per.append(bus_knowledge_per_obj.value)
        mark_potential_per.append(mark_potential_per_obj.value)
        fin_management_per.append(fin_management_per_obj.value)
        env_compliance_per.append(env_compliance_per_obj.value)
    
        
    data = {}
    data["pa_id"]=pa_id
    data["name_assessor"]=name_assessor
    data["date"]=date 
    data["start_time"]=start_time  
    data["end_time"]=end_time
    data["duration"]=duration
    data["type_assessed"]=type_assessed
    data["country"]=country
    data["code"]=code
    data["business_name"]=business_name
    data["legal_formation"]=legal_formation
    data["region"]=region
    data["zone"]=zone
    data["woreda"]=woreda
    data["telephone"]=telephone
    data["email"]=email
    data["contact_person"]=contact_person
    data["contact_designation"]=contact_designation
    data["contact_gender"]=contact_gender
    data["contact_educ"]=contact_educ
    data["hh_head_owner"]=hh_head_owner
    data["train_given_by"]=train_given_by
    data["train_satisfaction"]=train_satisfaction
    data["train_decision"]=train_decision
    data["have_license"]=have_license
    data["have_tin"]=have_tin
    data["year_estabilished"]=year_estabilished
    data["employment_male"]=employment_male 
    data["employment_female"]=employment_female 
    data["employment_total"]=employment_total
    data["avail_water"]=avail_water 
    data["avail_input_store"]=avail_input_store 
    data["avail_electricity "]=avail_electricity 
    data["avail_generator"]=avail_generator 
    data["avail_heating "]=avail_heating 
    data["ss_water_man"]=ss_water_man 
    data["ss_water_woman"]=ss_water_woman 
    data["ss_water_children"]=ss_water_children 
    data["ss_water_all"]=ss_water_all
    data["source_prot_inputs"]=source_prot_inputs
    data["get_business_advice"]=get_business_advice 
    data["access_poultry_mart"]=access_poultry_mart 
    data["agency_access_loan"]=agency_access_loan 
    data["agency_lender_type"]=agency_lender_type
    data["agency_amount"]=agency_amount 
    data["informal_access_loan"]=informal_access_loan 
    data["informal_lender_type"]=informal_lender_type 
    data["informal_amount"]=informal_amount
    data["finance_challenge1"]=finance_challenge1 
    data["finance_challenge2"]=finance_challenge2 
    data["finance_challenge3"]=finance_challenge3
    data["finance_challenge4"]=finance_challenge4
    data["finance_challenge_other"]=finance_challenge_other 
    
    data["level_educ"]=level_educ 
    data["experience_years"]=experience_years 
    data["received_training"]=received_training 
    data["employment_procedure"]=employment_procedure 
    data["provide_advisory_serv"]=provide_advisory_serv 
    data["convenient_poultry_house"]=convenient_poultry_house
    data["know_feed_types"]=know_feed_types
    data["source_of_water"]=source_of_water 
    data["quality_attribute"]=quality_attribute 
    data["poultry_disease"]=poultry_disease
    data["know_poultry_prtection"]=know_poultry_prtection 
    data["know_hygen"]=know_hygen 
    data["have_annual_plan"]=have_annual_plan 
    data["have_additional_income"]=have_additional_income  
    data["alternative_finance"]=alternative_finance 
    data["market_competition"]=market_competition 
    data["promotion"]=promotion 
    data["risk_price"]=risk_price 
    data["risk_ss_shortage"]=risk_ss_shortage
    data["working_capital"]=working_capital 
    data["financial_records"]=financial_records 
    data["risk_potential"]=risk_potential
    data["mitigation_method"]=mitigation_method 
    data["risk_health"]=risk_health
    
    data["prod_cycle2016"]=prod_cycle2016 
    data["prod_cycle2017"]=prod_cycle2017 
    data["prod_cycle2018"]=prod_cycle2018
    data["num_chicks_sold_farm2018"]=num_chicks_sold_farm2018 
    data["num_chicks_sold_gov_agent2018"]=num_chicks_sold_gov_agent2018
    data["num_chicks_sold_trader2018"]=num_chicks_sold_trader2018 
    data["num_chicks_sold_vill_agents2018"]=num_chicks_sold_vill_agents2018
    data["num_chicks_sold_other2018"]=num_chicks_sold_other2018 
    data["num_chicks_sold_total2018"]=num_chicks_sold_total2018
    data["num_chicks_sold_farm2017"]=num_chicks_sold_farm2017 
    data["num_chicks_sold_gov_agent2017"]=num_chicks_sold_gov_agent2017
    data["num_chicks_sold_trader2017"]=num_chicks_sold_trader2017 
    data["num_chicks_sold_vill_agents2017"]=num_chicks_sold_vill_agents2017
    data["num_chicks_sold_other2017"]=num_chicks_sold_other2017 
    data["num_chicks_sold_total2017"]=num_chicks_sold_total2017
    data["num_chicks_sold_farm2016"]=num_chicks_sold_farm2016 
    data["num_chicks_sold_gov_agent2016"]=num_chicks_sold_gov_agent2016
    data["num_chicks_sold_trader2016"]=num_chicks_sold_trader2016 
    data["num_chicks_sold_vill_agents2016"]=num_chicks_sold_vill_agents2016
    data["num_chicks_sold_other2016"]=num_chicks_sold_other2016 
    data["num_chicks_sold_total2016"]=num_chicks_sold_total2016
    data["turn_over_farm2018"]=turn_over_farm2018 
    data["turn_over_gov_agent2018"]=turn_over_gov_agent2018
    data["turn_over_trader2018"]=turn_over_trader2018 
    data["turn_over_vill_agents2018"]=turn_over_vill_agents2018
    data["turn_over_other2018"]=turn_over_other2018 
    data["turn_over_total2018"]=turn_over_total2018
    data["turn_over_farm2017"]=turn_over_farm2017 
    data["turn_over_gov_agent2017"]=turn_over_gov_agent2017
    data["turn_over_trader2017"]=turn_over_trader2017 
    data["turn_over_vill_agents2017"]=turn_over_vill_agents2017
    data["turn_over_other2017"]=turn_over_other2017 
    data["turn_over_total2017"]=turn_over_total2017
    data["turn_over_farm2016"]=turn_over_farm2016 
    data["turn_over_gov_agent2016"]=turn_over_gov_agent2016
    data["turn_over_trader2016"]=turn_over_trader2016
    data["turn_over_vill_agents2016"]=turn_over_vill_agents2016
    data["turn_over_other2016"]=turn_over_other2016
    data["turn_over_total2016"]=turn_over_total2016 
    data["sold_to_unique_farmers2018"]=sold_to_unique_farmers2018
    data["sold_to_unique_farmers2017"]=sold_to_unique_farmers2017
    data["sold_to_unique_farmers2016"]=sold_to_unique_farmers2016
    
    data["mgmt_capability_lev"]=mgmt_capability_lev
    data["inst_competency_lev"]=inst_competency_lev
    data["bus_knowledge_lev"]=bus_knowledge_lev
    data["mark_potential_lev"]=mark_potential_lev
    data["fin_management_lev"]=fin_management_lev
    data["env_compliance_lev"]=env_compliance_lev
    data["mgmt_capability_lev"]=mgmt_capability_lev
    data["inst_competency_lev"]=inst_competency_lev
    data["bus_knowledge_lev"]=bus_knowledge_lev
    data["mark_potential_lev"]=mark_potential_lev
    data["fin_management_lev"]=fin_management_lev
    data["env_compliance_lev"]=env_compliance_lev
    df = pd.DataFrame.from_dict(data)
    
    return df



def get_data2(file_path, files):
    
    [
    pa_id,
    name_assessor,
    date,
    start_time,  
    end_time,
    duration,
    type_assessed,
    country,
    code,
    business_name,
    legal_formation,
    region,
    zone,
    woreda,
    telephone,
    email,
    contact_person,
    contact_designation,
    contact_gender,
    contact_educ,
    hh_head_owner,
    train_given_by,
    train_satisfaction,
    train_decision,
    have_license,
    have_tin,
    year_estabilished,
    employment_male, 
    employment_female, 
    employment_total,
    avail_water, 
    avail_input_store, 
    avail_electricity, 
    avail_generator, 
    avail_heating, 
    ss_water_man, 
    ss_water_woman, 
    ss_water_children, 
    ss_water_all,
    source_prot_inputs,
    get_business_advice, 
    access_poultry_mart, 
    agency_access_loan, 
    agency_lender_type,
    agency_amount, 
    informal_access_loan, 
    informal_lender_type, 
    informal_amount,
    finance_challenge1, 
    finance_challenge2, 
    finance_challenge3,
    finance_challenge4,
    finance_challenge_other, 
    
    level_educ, 
    experience_years, 
    received_training, 
    employment_procedure, 
    provide_advisory_serv, 
    convenient_poultry_house,
    know_feed_types,
    source_of_water, 
    quality_attribute, 
    poultry_disease,
    know_poultry_prtection, 
    know_hygen, 
    have_annual_plan, 
    have_additional_income,  
    alternative_finance, 
    market_competition, 
    promotion, 
    risk_price, 
    risk_ss_shortage,
    working_capital, 
    financial_records, 
    risk_potential,
    mitigation_method, 
    risk_health,
    
    prod_cycle2016, 
    prod_cycle2017, 
    prod_cycle2018,
    num_chicks_sold_farm2018, 
    num_chicks_sold_gov_agent2018,
    num_chicks_sold_trader2018, 
    num_chicks_sold_vill_agents2018,
    num_chicks_sold_other2018, 
    num_chicks_sold_total2018,
    num_chicks_sold_farm2017, 
    num_chicks_sold_gov_agent2017,
    num_chicks_sold_trader2017, 
    num_chicks_sold_vill_agents2017,
    num_chicks_sold_other2017, 
    num_chicks_sold_total2017,
    num_chicks_sold_farm2016, 
    num_chicks_sold_gov_agent2016,
    num_chicks_sold_trader2016, 
    num_chicks_sold_vill_agents2016,
    num_chicks_sold_other2016, 
    num_chicks_sold_total2016,
    turn_over_farm2018, 
    turn_over_gov_agent2018,
    turn_over_trader2018, 
    turn_over_vill_agents2018,
    turn_over_other2018, 
    turn_over_total2018,
    turn_over_farm2017, 
    turn_over_gov_agent2017,
    turn_over_trader2017, 
    turn_over_vill_agents2017,
    turn_over_other2017, 
    turn_over_total2017,
    turn_over_farm2016, 
    turn_over_gov_agent2016,
    turn_over_trader2016,
    turn_over_vill_agents2016,
    turn_over_other2016,
    turn_over_total2016, 
    sold_to_unique_farmers2018,
    sold_to_unique_farmers2017,
    sold_to_unique_farmers2016,
    
    mgmt_capability_lev,
    inst_competency_lev,
    bus_knowledge_lev,
    mark_potential_lev,
    fin_management_lev,
    env_compliance_lev,
    mgmt_capability_lev,
    inst_competency_lev,
    bus_knowledge_lev,
    mark_potential_lev,
    fin_management_lev,
    env_compliance_lev
    ]= ([] for i in range(131))
    
    for file in files:
        match = re.search(r'\d*', file)
        pa_id.append(match.group())
        
        workbook = load_workbook(file_path+file, data_only=True)
                    
        worksheet = workbook['BIETH']
        print(f'Working on file: {file} ...')
        print('Sheet: BIETH ...')
    
        name_assessor_obj = worksheet['G7']
        date_obj  = worksheet['A11']
        start_time_obj  = worksheet['B11']
        end_time_obj = worksheet['C11']
        duration_obj = worksheet['D11']
        type_assessed_obj = worksheet['F10']
        country_obj = worksheet['G10']
        code_obj = worksheet['H10']
        business_name_obj = worksheet['C16']
        legal_formation_obj = worksheet['C17']
        region_obj = worksheet['C18']
        zone_obj = worksheet['C19']
        woreda_obj = worksheet['C20']
        telephone_obj = worksheet['C21']
        email_obj = worksheet['C22']
        
        contact_person_obj = worksheet['G16']
        contact_designation_obj = worksheet['H16']
        contact_gender_obj = worksheet['G17']
        contact_educ_obj = worksheet['G18']
        hh_head_owner_obj = worksheet['G20']
        train_given_by_obj = worksheet['C25']
        train_satisfaction_obj = worksheet['F25']
        train_decision_obj = worksheet['G25']
        
        have_license_obj = worksheet['C31']
        have_tin_obj = worksheet['C32']
        
        year_estabilished_obj = worksheet['G31']
        
        employment_male_obj = worksheet['A37']
        employment_female_obj = worksheet['B37']
        employment_total_obj = worksheet['C37']
        
        avail_water_obj = worksheet['H34']
        avail_input_store_obj = worksheet['H35']
        avail_electricity_obj = worksheet['H36']
        avail_generator_obj = worksheet['H37']
        avail_heating_obj = worksheet['H38']
        
        ss_water_man_obj = worksheet['C41']
        ss_water_woman_obj = worksheet['F41']
        ss_water_children_obj = worksheet['G41']
        ss_water_all_obj = worksheet['H41']
        
        source_prot_inputs_obj=worksheet['A44']
        get_business_advice_obj = worksheet['C44']
        access_poultry_mart_obj = worksheet['F44']
        
        agency_access_loan_obj = worksheet['C46']
        agency_lender_type_obj = worksheet['F46']
        agency_amount_obj = worksheet['H46']
        informal_access_loan_obj = worksheet['C47']
        informal_lender_type_obj = worksheet['F47'] 
        informal_amount_obj = worksheet['H47']
        
        finance_challenge1_obj = worksheet['A50']
        finance_challenge2_obj = worksheet['C50']
        finance_challenge3_obj = worksheet['E50']
        finance_challenge4_obj = worksheet['G50']
        finance_challenge_other_obj = worksheet['D51']
    
        worksheet = workbook['QETH']
        print(f'Working on file: {file} ...')
        print('Sheet: QETH ...')
    
        level_educ_obj = worksheet['F14']
        experience_years_obj = worksheet['F20']
        received_training_obj = worksheet['F26']
        employment_procedure_obj = worksheet['F33']
        provide_advisory_serv_obj = worksheet['F39']
        convenient_poultry_house_obj = worksheet['F49']
        know_feed_types_obj = worksheet['F55']    
        source_of_water_obj = worksheet['F61']
        quality_attribute_obj = worksheet['F67']
        poultry_disease_obj = worksheet['F73']
        know_poultry_prtection_obj = worksheet['F79']
        know_hygen_obj = worksheet['F86']
        have_annual_plan_obj = worksheet['F96']
        have_additional_income_obj = worksheet['F103']
        alternative_finance_obj = worksheet['F110']
        market_competition_obj = worksheet['F120']
        promotion_obj = worksheet['F126']
        risk_price_obj = worksheet['F133']
        risk_ss_shortage_obj = worksheet['F139']
        working_capital_obj = worksheet['F149']
        financial_records_obj = worksheet['F155']
        risk_potential_obj = worksheet['F165']
        mitigation_method_obj = worksheet['F171']
        risk_health_obj = worksheet['F178']
    
        worksheet = workbook['Additional Info']
        print(f'Working on file: {file} ...')
        print('Sheet: Additional Info ...')
    
        prod_cycle2016_obj = worksheet['C15']
        prod_cycle2017_obj = worksheet['D15']
        prod_cycle2018_obj = worksheet['E15']
        
        num_chicks_sold_farm2018_obj = worksheet['C20']
        num_chicks_sold_gov_agent2018_obj = worksheet['D20']
        num_chicks_sold_trader2018_obj = worksheet['E20']
        num_chicks_sold_vill_agents2018_obj = worksheet['F20']
        num_chicks_sold_other2018_obj = worksheet['G20']
        num_chicks_sold_total2018_obj = worksheet['H20']
        
        turn_over_farm2018_obj = worksheet['C21']
        turn_over_gov_agent2018_obj = worksheet['D21']
        turn_over_trader2018_obj = worksheet['E21']
        turn_over_vill_agents2018_obj = worksheet['F21']
        turn_over_other2018_obj = worksheet['G21']
        turn_over_total2018_obj = worksheet['H21']
        
        sold_to_unique_farmers2018_obj = worksheet['G22']
    
        num_chicks_sold_farm2017_obj = worksheet['C26']
        num_chicks_sold_gov_agent2017_obj = worksheet['D26']
        num_chicks_sold_trader2017_obj = worksheet['E26']
        num_chicks_sold_vill_agents2017_obj = worksheet['F26']
        num_chicks_sold_other2017_obj = worksheet['G26']
        num_chicks_sold_total2017_obj = worksheet['H26']
        
        turn_over_farm2017_obj = worksheet['C27']
        turn_over_gov_agent2017_obj = worksheet['D27']
        turn_over_trader2017_obj = worksheet['E27']
        turn_over_vill_agents2017_obj = worksheet['F27']
        turn_over_other2017_obj = worksheet['G27']
        turn_over_total2017_obj = worksheet['H27']
        
        sold_to_unique_farmers2017_obj = worksheet['G28']
    
        num_chicks_sold_farm2016_obj = worksheet['C32']
        num_chicks_sold_gov_agent2016_obj = worksheet['D32']
        num_chicks_sold_trader2016_obj = worksheet['E32']
        num_chicks_sold_vill_agents2016_obj = worksheet['F32']
        num_chicks_sold_other2016_obj = worksheet['G32']
        num_chicks_sold_total2016_obj = worksheet['H32']
                
        turn_over_farm2016_obj = worksheet['C33']
        turn_over_gov_agent2016_obj = worksheet['D33']
        turn_over_trader2016_obj = worksheet['E33']
        turn_over_vill_agents2016_obj = worksheet['F33']
        turn_over_other2016_obj = worksheet['G33']
        turn_over_total2016_obj =  worksheet['H33']
    
        sold_to_unique_farmers2016_obj = worksheet['G34']
             
        
        worksheet = workbook['RETH-EN']
        print(f'Working on file: {file} ...')
        print('Sheet: RETH-EN ...')
        
        mgmt_capability_lev_obj=worksheet['D15']
        inst_competency_lev_obj=worksheet['D16']
        bus_knowledge_lev_obj=worksheet['D17']
        mark_potential_lev_obj=worksheet['D18']
        fin_management_lev_obj=worksheet['D19']
        env_compliance_lev_obj=worksheet['D20']
        mgmt_capability_lev_obj=worksheet['E15']
        inst_competency_lev_obj=worksheet['E16']
        bus_knowledge_lev_obj=worksheet['E17']
        mark_potential_lev_obj=worksheet['E18']
        fin_management_lev_obj=worksheet['E19']
        env_compliance_lev_obj=worksheet['E20']
        
        name_assessor.append(name_assessor_obj.value)
        date.append(date_obj.value)
        start_time.append(start_time_obj.value)
        end_time.append(end_time_obj.value)
        duration.append(duration_obj.value)
        type_assessed.append(type_assessed_obj.value)
        country.append(country_obj.value)
        code.append(code_obj.value)
        business_name.append(business_name_obj.value)
        legal_formation.append(legal_formation_obj.value)
        region.append(region_obj.value)
        zone.append(zone_obj.value)
        woreda.append(woreda_obj.value)
        telephone.append(telephone_obj.value)
        email.append(email_obj.value)
        contact_person.append(contact_person_obj.value)
        contact_designation.append(contact_designation_obj.value)
        contact_gender.append(contact_gender_obj.value)
        contact_educ.append(contact_educ_obj.value)
        hh_head_owner.append(hh_head_owner_obj.value)
        train_given_by.append(train_given_by_obj.value)
        train_satisfaction.append(train_satisfaction_obj.value)
        train_decision.append(train_decision_obj.value)
        have_license.append(have_license_obj.value)
        have_tin.append(have_tin_obj.value)
        year_estabilished.append(year_estabilished_obj.value)
        employment_male.append(employment_male_obj.value)
        employment_female.append(employment_female_obj.value)
        employment_total.append(employment_total_obj.value)
        avail_water.append(avail_water_obj.value)
        avail_input_store.append(avail_input_store_obj.value)
        avail_electricity.append(avail_electricity_obj.value)
        avail_generator.append(avail_generator_obj.value)
        avail_heating.append(avail_heating_obj.value)
        ss_water_man.append(ss_water_man_obj.value)
        ss_water_woman.append(ss_water_woman_obj.value)
        ss_water_children.append(ss_water_children_obj.value)
        ss_water_all.append(ss_water_all_obj.value)
        source_prot_inputs.append(source_prot_inputs_obj.value)
        get_business_advice.append(get_business_advice_obj.value)
        access_poultry_mart.append(access_poultry_mart_obj.value)
        agency_access_loan.append(agency_access_loan_obj.value)
        agency_lender_type.append(agency_lender_type_obj.value)
        agency_amount.append(agency_amount_obj.value)
        informal_access_loan.append(informal_access_loan_obj.value)
        informal_lender_type.append(informal_lender_type_obj.value)
        informal_amount.append(informal_amount_obj.value)
        finance_challenge1.append(finance_challenge1_obj.value)
        finance_challenge2.append(finance_challenge2_obj.value)
        finance_challenge3.append(finance_challenge3_obj.value)
        finance_challenge4.append(finance_challenge4_obj.value)
        finance_challenge_other.append(finance_challenge_other_obj.value)
        
        level_educ.append(level_educ_obj.value)
        experience_years.append(experience_years_obj.value)
        received_training.append(received_training_obj.value)
        employment_procedure.append(employment_procedure_obj.value)
        provide_advisory_serv.append(provide_advisory_serv_obj.value)
        convenient_poultry_house.append(convenient_poultry_house_obj.value)
        know_feed_types.append(know_feed_types_obj.value)
        source_of_water.append(source_of_water_obj.value)
        quality_attribute.append(quality_attribute_obj.value)
        poultry_disease.append(poultry_disease_obj.value)
        know_poultry_prtection.append(know_poultry_prtection_obj.value)
        know_hygen.append(know_hygen_obj.value)
        have_annual_plan.append(have_annual_plan_obj.value)
        have_additional_income.append(have_additional_income_obj.value)
        alternative_finance.append(alternative_finance_obj.value)
        market_competition.append(market_competition_obj.value)
        promotion.append(promotion_obj.value)
        risk_price.append(risk_price_obj.value)
        risk_ss_shortage.append(risk_ss_shortage_obj.value)
        working_capital.append(working_capital_obj.value)
        financial_records.append(financial_records_obj.value)
        risk_potential.append(risk_potential_obj.value)
        mitigation_method.append(mitigation_method_obj.value)
        risk_health.append(risk_health_obj.value)
        
        prod_cycle2016.append(prod_cycle2016_obj.value)
        prod_cycle2017.append(prod_cycle2017_obj.value)
        prod_cycle2018.append(prod_cycle2018_obj.value)
        num_chicks_sold_farm2018.append(num_chicks_sold_farm2018_obj.value)
        num_chicks_sold_gov_agent2018.append(num_chicks_sold_gov_agent2018_obj.value)
        num_chicks_sold_trader2018.append(num_chicks_sold_trader2018_obj.value)
        num_chicks_sold_vill_agents2018.append(num_chicks_sold_vill_agents2018_obj.value)
        num_chicks_sold_other2018.append(num_chicks_sold_other2018_obj.value)
        num_chicks_sold_total2018.append(num_chicks_sold_total2018_obj.value)
        num_chicks_sold_farm2017.append(num_chicks_sold_farm2017_obj.value)
        num_chicks_sold_gov_agent2017.append(num_chicks_sold_gov_agent2017_obj.value)
        num_chicks_sold_trader2017.append(num_chicks_sold_trader2017_obj.value)
        num_chicks_sold_vill_agents2017.append(num_chicks_sold_vill_agents2017_obj.value)
        num_chicks_sold_other2017.append(num_chicks_sold_other2017_obj.value)
        num_chicks_sold_total2017.append(num_chicks_sold_total2017_obj.value)
        num_chicks_sold_farm2016.append(num_chicks_sold_farm2016_obj.value)
        num_chicks_sold_gov_agent2016.append(num_chicks_sold_gov_agent2016_obj.value)
        num_chicks_sold_trader2016.append(num_chicks_sold_trader2016_obj.value)
        num_chicks_sold_vill_agents2016.append(num_chicks_sold_vill_agents2016_obj.value)
        num_chicks_sold_other2016.append(num_chicks_sold_other2016_obj.value)
        num_chicks_sold_total2016.append(num_chicks_sold_total2016_obj.value)
        turn_over_farm2018.append(turn_over_farm2018_obj.value)
        turn_over_gov_agent2018.append(turn_over_gov_agent2018_obj.value)
        turn_over_trader2018.append(turn_over_trader2018_obj.value)
        turn_over_vill_agents2018.append(turn_over_vill_agents2018_obj.value)
        turn_over_other2018.append(turn_over_other2018_obj.value)
        turn_over_total2018.append(turn_over_total2018_obj.value)
        turn_over_farm2017.append(turn_over_farm2017_obj.value)
        turn_over_gov_agent2017.append(turn_over_gov_agent2017_obj.value)
        turn_over_trader2017.append(turn_over_trader2017_obj.value)
        turn_over_vill_agents2017.append(turn_over_vill_agents2017_obj.value)
        turn_over_other2017.append(turn_over_other2017_obj.value)
        turn_over_total2017.append(turn_over_total2017_obj.value)
        turn_over_farm2016.append(turn_over_farm2016_obj.value)
        turn_over_gov_agent2016.append(turn_over_gov_agent2016_obj.value)
        turn_over_trader2016.append(turn_over_trader2016_obj.value)
        turn_over_vill_agents2016.append(turn_over_vill_agents2016_obj.value)
        turn_over_other2016.append(turn_over_other2016_obj.value)
        turn_over_total2016.append(turn_over_total2016_obj.value)
        sold_to_unique_farmers2018.append(sold_to_unique_farmers2018_obj.value)
        sold_to_unique_farmers2017.append(sold_to_unique_farmers2017_obj.value)
        sold_to_unique_farmers2016.append(sold_to_unique_farmers2016_obj.value)
    
        mgmt_capability_lev.append(mgmt_capability_lev_obj.value)
        inst_competency_lev.append(inst_competency_lev_obj.value)
        bus_knowledge_lev.append(bus_knowledge_lev_obj.value)
        mark_potential_lev.append(mark_potential_lev_obj.value)
        fin_management_lev.append(fin_management_lev_obj.value)
        env_compliance_lev.append(env_compliance_lev_obj.value)
        mgmt_capability_lev.append(mgmt_capability_lev_obj.value)
        inst_competency_lev.append(inst_competency_lev_obj.value)
        bus_knowledge_lev.append(bus_knowledge_lev_obj.value)
        mark_potential_lev.append(mark_potential_lev_obj.value)
        fin_management_lev.append(fin_management_lev_obj.value)
        env_compliance_lev.append(env_compliance_lev_obj.value)
    
        
    data = {}
    data["pa_id"]=pa_id
    data["name_assessor"]=name_assessor
    data["date"]=date 
    data["start_time"]=start_time  
    data["end_time"]=end_time
    data["duration"]=duration
    data["type_assessed"]=type_assessed
    data["country"]=country
    data["code"]=code
    data["business_name"]=business_name
    data["legal_formation"]=legal_formation
    data["region"]=region
    data["zone"]=zone
    data["woreda"]=woreda
    data["telephone"]=telephone
    data["email"]=email
    data["contact_person"]=contact_person
    data["contact_designation"]=contact_designation
    data["contact_gender"]=contact_gender
    data["contact_educ"]=contact_educ
    data["hh_head_owner"]=hh_head_owner
    data["train_given_by"]=train_given_by
    data["train_satisfaction"]=train_satisfaction
    data["train_decision"]=train_decision
    data["have_license"]=have_license
    data["have_tin"]=have_tin
    data["year_estabilished"]=year_estabilished
    data["employment_male"]=employment_male 
    data["employment_female"]=employment_female 
    data["employment_total"]=employment_total
    data["avail_water"]=avail_water 
    data["avail_input_store"]=avail_input_store 
    data["avail_electricity "]=avail_electricity 
    data["avail_generator"]=avail_generator 
    data["avail_heating "]=avail_heating 
    data["ss_water_man"]=ss_water_man 
    data["ss_water_woman"]=ss_water_woman 
    data["ss_water_children"]=ss_water_children 
    data["ss_water_all"]=ss_water_all
    data["source_prot_inputs"]=source_prot_inputs
    data["get_business_advice"]=get_business_advice 
    data["access_poultry_mart"]=access_poultry_mart 
    data["agency_access_loan"]=agency_access_loan 
    data["agency_lender_type"]=agency_lender_type
    data["agency_amount"]=agency_amount 
    data["informal_access_loan"]=informal_access_loan 
    data["informal_lender_type"]=informal_lender_type 
    data["informal_amount"]=informal_amount
    data["finance_challenge1"]=finance_challenge1 
    data["finance_challenge2"]=finance_challenge2 
    data["finance_challenge3"]=finance_challenge3
    data["finance_challenge4"]=finance_challenge4
    data["finance_challenge_other"]=finance_challenge_other 
    
    data["level_educ"]=level_educ 
    data["experience_years"]=experience_years 
    data["received_training"]=received_training 
    data["employment_procedure"]=employment_procedure 
    data["provide_advisory_serv"]=provide_advisory_serv 
    data["convenient_poultry_house"]=convenient_poultry_house
    data["know_feed_types"]=know_feed_types
    data["source_of_water"]=source_of_water 
    data["quality_attribute"]=quality_attribute 
    data["poultry_disease"]=poultry_disease
    data["know_poultry_prtection"]=know_poultry_prtection 
    data["know_hygen"]=know_hygen 
    data["have_annual_plan"]=have_annual_plan 
    data["have_additional_income"]=have_additional_income  
    data["alternative_finance"]=alternative_finance 
    data["market_competition"]=market_competition 
    data["promotion"]=promotion 
    data["risk_price"]=risk_price 
    data["risk_ss_shortage"]=risk_ss_shortage
    data["working_capital"]=working_capital 
    data["financial_records"]=financial_records 
    data["risk_potential"]=risk_potential
    data["mitigation_method"]=mitigation_method 
    data["risk_health"]=risk_health
    
    data["prod_cycle2016"]=prod_cycle2016 
    data["prod_cycle2017"]=prod_cycle2017 
    data["prod_cycle2018"]=prod_cycle2018
    data["num_chicks_sold_farm2018"]=num_chicks_sold_farm2018 
    data["num_chicks_sold_gov_agent2018"]=num_chicks_sold_gov_agent2018
    data["num_chicks_sold_trader2018"]=num_chicks_sold_trader2018 
    data["num_chicks_sold_vill_agents2018"]=num_chicks_sold_vill_agents2018
    data["num_chicks_sold_other2018"]=num_chicks_sold_other2018 
    data["num_chicks_sold_total2018"]=num_chicks_sold_total2018
    data["num_chicks_sold_farm2017"]=num_chicks_sold_farm2017 
    data["num_chicks_sold_gov_agent2017"]=num_chicks_sold_gov_agent2017
    data["num_chicks_sold_trader2017"]=num_chicks_sold_trader2017 
    data["num_chicks_sold_vill_agents2017"]=num_chicks_sold_vill_agents2017
    data["num_chicks_sold_other2017"]=num_chicks_sold_other2017 
    data["num_chicks_sold_total2017"]=num_chicks_sold_total2017
    data["num_chicks_sold_farm2016"]=num_chicks_sold_farm2016 
    data["num_chicks_sold_gov_agent2016"]=num_chicks_sold_gov_agent2016
    data["num_chicks_sold_trader2016"]=num_chicks_sold_trader2016 
    data["num_chicks_sold_vill_agents2016"]=num_chicks_sold_vill_agents2016
    data["num_chicks_sold_other2016"]=num_chicks_sold_other2016 
    data["num_chicks_sold_total2016"]=num_chicks_sold_total2016
    data["turn_over_farm2018"]=turn_over_farm2018 
    data["turn_over_gov_agent2018"]=turn_over_gov_agent2018
    data["turn_over_trader2018"]=turn_over_trader2018 
    data["turn_over_vill_agents2018"]=turn_over_vill_agents2018
    data["turn_over_other2018"]=turn_over_other2018 
    data["turn_over_total2018"]=turn_over_total2018
    data["turn_over_farm2017"]=turn_over_farm2017 
    data["turn_over_gov_agent2017"]=turn_over_gov_agent2017
    data["turn_over_trader2017"]=turn_over_trader2017 
    data["turn_over_vill_agents2017"]=turn_over_vill_agents2017
    data["turn_over_other2017"]=turn_over_other2017 
    data["turn_over_total2017"]=turn_over_total2017
    data["turn_over_farm2016"]=turn_over_farm2016 
    data["turn_over_gov_agent2016"]=turn_over_gov_agent2016
    data["turn_over_trader2016"]=turn_over_trader2016
    data["turn_over_vill_agents2016"]=turn_over_vill_agents2016
    data["turn_over_other2016"]=turn_over_other2016
    data["turn_over_total2016"]=turn_over_total2016 
    data["sold_to_unique_farmers2018"]=sold_to_unique_farmers2018
    data["sold_to_unique_farmers2017"]=sold_to_unique_farmers2017
    data["sold_to_unique_farmers2016"]=sold_to_unique_farmers2016
    
    data["mgmt_capability_lev"]=mgmt_capability_lev
    data["inst_competency_lev"]=inst_competency_lev
    data["bus_knowledge_lev"]=bus_knowledge_lev
    data["mark_potential_lev"]=mark_potential_lev
    data["fin_management_lev"]=fin_management_lev
    data["env_compliance_lev"]=env_compliance_lev
    data["mgmt_capability_lev"]=mgmt_capability_lev
    data["inst_competency_lev"]=inst_competency_lev
    data["bus_knowledge_lev"]=bus_knowledge_lev
    data["mark_potential_lev"]=mark_potential_lev
    data["fin_management_lev"]=fin_management_lev
    data["env_compliance_lev"]=env_compliance_lev
    df = pd.DataFrame.from_dict(data)
    
    return df



def get_data_fd1(file_path, files):
    [
    pa_id,
    name_assessor,
    date,
    start_time,  
    end_time,
    duration,
    type_assessed,
    country,
    code,
    business_name,
    legal_formation,
    region,
    zone,
    woreda,
    telephone,
    email,
    contact_person,
    contact_designation,
    contact_gender,
    contact_educ,
    hh_head_owner,
    train_given_by,
    train_satisfaction,
    train_decision,
    have_license,
    have_tin,
    year_estabilished,
    employment_male, 
    employment_female, 
    employment_total,
    avail_water, 
    avail_input_store, 
    avail_electricity,  
    get_business_advice, 
    access_poultry_mart, 
    agency_access_loan, 
    agency_lender_type,
    agency_amount, 
    informal_access_loan, 
    informal_lender_type, 
    informal_amount,
    finance_challenge1, 
    finance_challenge2, 
    finance_challenge3,
    finance_challenge4,
    finance_challenge_other, 
    level_educ, 
    experience_years, 
    received_training, 
    employment_procedure, 
    provide_advisory_serv, 
    know_feed_types,
    know_feed_handling,
    feed_store_standard,
    store_sanitation,
    quality_control,
    have_annual_plan, 
    have_additional_income,  
    alternative_finance, 
    market_competition, 
    promotion, 
    risk_price, 
    risk_ss_shortage,
    working_capital, 
    financial_records, 
    risk_potential,
    mitigation_method, 
    risk_health,
    num_purchase_inputs2018, 
    num_purchase_inputs2017, 
    num_purchase_inputs2016,
    inputs_sold_gk_farmer2018,
    inputs_sold_gk_gov_agent2018,
    inputs_sold_gk_poult_agent2018,
    inputs_sold_gk_vill_agent2018,
    inputs_sold_gk_other2018,
    inputs_sold_gk_total2018,
    sales_turnover_farmer2018,
    sales_turnover_gov_agent2018,
    sales_turnover_poult_agent2018,
    sales_turnover_vill_agent2018,
    sales_turnover_other2018,
    sales_turnover_total2018,
    inputs_sold_gk_farmer2017,
    inputs_sold_gk_gov_agent2017,
    inputs_sold_gk_poult_agent2017,
    inputs_sold_gk_vill_agent2017,
    inputs_sold_gk_other2017,
    inputs_sold_gk_total2017,
    sales_turnover_farmer2017,
    sales_turnover_gov_agent2017,
    sales_turnover_poult_agent2017,
    sales_turnover_vill_agent2017,
    sales_turnover_other2017,
    sales_turnover_total2017,
    inputs_sold_gk_farmer2016,
    inputs_sold_gk_gov_agent2016,
    inputs_sold_gk_poult_agent2016,
    inputs_sold_gk_vill_agent2016,
    inputs_sold_gk_other2016,
    inputs_sold_gk_total2016,
    sales_turnover_farmer2016,
    sales_turnover_gov_agent2016,
    sales_turnover_poult_agent2016,
    sales_turnover_vill_agent2016,
    sales_turnover_other2016,
    sales_turnover_total2016,
    num_unique_farmer2018,
    num_unique_farmer2017,
    num_unique_farmer2016,
    
    mgmt_capability_lev,
    inst_competency_lev,
    bus_knowledge_lev,
    mark_potential_lev,
    fin_management_lev,
    env_compliance_lev,
    mgmt_capability_lev,
    inst_competency_lev,
    bus_knowledge_lev,
    mark_potential_lev,
    fin_management_lev,
    env_compliance_lev
    ]= ([] for i in range(122))
    
    
    for file in files:
        match = re.search(r'\d*', file)
        pa_id.append(match.group())
        
        workbook = load_workbook(file_path+file, data_only=True)
        
        worksheet = workbook['BIETH']
        print(f'Working on file: {file} ...')
        print('Sheet: BIETH ...')
         
        name_assessor_obj=worksheet['G6']
        date_obj=worksheet['A10']
        start_time_obj=worksheet['B10']
        end_time_obj=worksheet['C10']
        duration_obj=worksheet['D10']
        
        type_assessed_obj=worksheet['F9']
        country_obj=worksheet['G9']
        code_obj=worksheet['H9']
        
        business_name_obj=worksheet['C15']
        legal_formation_obj=worksheet['C16']
        region_obj=worksheet['C17']
        zone_obj=worksheet['C18']
        woreda_obj=worksheet['C19']
        telephone_obj=worksheet['C20']
        email_obj=worksheet['C21']
        
        contact_person_obj=worksheet['G15']
        contact_designation_obj=worksheet['H15']
        contact_gender_obj=worksheet['G16']
        contact_educ_obj=worksheet['G17']
        hh_head_owner_obj=worksheet['G19']
        
        train_given_by_obj=worksheet['C24']
        train_satisfaction_obj=worksheet['F24']
        train_decision_obj=worksheet['G24']
        
        have_license_obj=worksheet['C30']
        have_tin_obj=worksheet['C31']
        year_estabilished_obj=worksheet['G30']
        
        employment_male_obj=worksheet['A35']
        employment_female_obj=worksheet['B35']
        employment_total_obj=worksheet['C35']
        avail_water_obj=worksheet['H33']
        avail_input_store_obj=worksheet['H34']
        avail_electricity_obj=worksheet['H35']
        
        get_business_advice_obj=worksheet['A39']
        access_poultry_mart_obj=worksheet['F39']
        
        agency_access_loan_obj=worksheet['C41']
        agency_lender_type_obj=worksheet['F41']
        agency_amount_obj=worksheet['H41']
        informal_access_loan_obj=worksheet['C42']
        informal_lender_type_obj=worksheet['F42']
        informal_amount_obj=worksheet['H42']
        
        finance_challenge1_obj=worksheet['A45']
        finance_challenge2_obj=worksheet['C45']
        finance_challenge3_obj=worksheet['E45']
        finance_challenge4_obj=worksheet['G45']
        finance_challenge_other_obj=worksheet['C46']
        
        worksheet = workbook['QETH']
        print(f'Working on file: {file} ...')
        print('Sheet: QETH ...')
    
        level_educ_obj=worksheet['F14']
        experience_years_obj=worksheet['F20']
        received_training_obj=worksheet['F26']
        employment_procedure_obj=worksheet['F33']
        provide_advisory_serv_obj=worksheet['F39']
        
        know_feed_types_obj=worksheet['F49']
        know_feed_handling_obj=worksheet['F55']
        feed_store_standard_obj=worksheet['F62']
        store_sanitation_obj=worksheet['F68']
        quality_control_obj=worksheet['F75']
        have_annual_plan_obj=worksheet['F85']
        have_additional_income_obj=worksheet['F92']
        alternative_finance_obj=worksheet['F99']
        market_competition_obj=worksheet['F109']
        promotion_obj=worksheet['F115']
        risk_price_obj=worksheet['F122']
        risk_ss_shortage_obj=worksheet['F128']
        working_capital_obj=worksheet['F138']
        financial_records_obj=worksheet['F144']
        risk_potential_obj=worksheet['F154']
        mitigation_method_obj=worksheet['F160']
        risk_health_obj=worksheet['F167']
        
        worksheet = workbook['Additional Info']
        print(f'Working on file: {file} ...')
        print('Sheet: Additional Info ...')
        
        num_purchase_inputs2018_obj=worksheet['C15']
        num_purchase_inputs2017_obj=worksheet['D15']
        num_purchase_inputs2016_obj=worksheet['E15']    
        
        inputs_sold_gk_farmer2018_obj=worksheet['C20']
        inputs_sold_gk_gov_agent2018_obj=worksheet['D20']
        inputs_sold_gk_poult_agent2018_obj=worksheet['E20']
        inputs_sold_gk_vill_agent2018_obj=worksheet['F20']
        inputs_sold_gk_other2018_obj=worksheet['G20']
        inputs_sold_gk_total2018_obj=worksheet['H20']    
        
        sales_turnover_farmer2018_obj=worksheet['C21']
        sales_turnover_gov_agent2018_obj=worksheet['D21']
        sales_turnover_poult_agent2018_obj=worksheet['E21']
        sales_turnover_vill_agent2018_obj=worksheet['F21']
        sales_turnover_other2018_obj=worksheet['G21']
        sales_turnover_total2018_obj=worksheet['H21']    
    
        num_unique_farmer2018_obj=worksheet['G22']
        
        inputs_sold_gk_farmer2017_obj=worksheet['C26']
        inputs_sold_gk_gov_agent2017_obj=worksheet['D26']
        inputs_sold_gk_poult_agent2017_obj=worksheet['E26']
        inputs_sold_gk_vill_agent2017_obj=worksheet['F26']
        inputs_sold_gk_other2017_obj=worksheet['G26']
        inputs_sold_gk_total2017_obj=worksheet['H26']    
        
        sales_turnover_farmer2017_obj=worksheet['C27']
        sales_turnover_gov_agent2017_obj=worksheet['D27']
        sales_turnover_poult_agent2017_obj=worksheet['E27']
        sales_turnover_vill_agent2017_obj=worksheet['F27']
        sales_turnover_other2017_obj=worksheet['G27']
        sales_turnover_total2017_obj=worksheet['H27']    
    
        num_unique_farmer2017_obj=worksheet['G28']
        
        inputs_sold_gk_farmer2016_obj=worksheet['C32']
        inputs_sold_gk_gov_agent2016_obj=worksheet['D32']
        inputs_sold_gk_poult_agent2016_obj=worksheet['E32']
        inputs_sold_gk_vill_agent2016_obj=worksheet['F32']
        inputs_sold_gk_other2016_obj=worksheet['G32']
        inputs_sold_gk_total2016_obj=worksheet['H32'] 
        
        sales_turnover_farmer2016_obj=worksheet['C33']
        sales_turnover_gov_agent2016_obj=worksheet['D33']
        sales_turnover_poult_agent2016_obj=worksheet['E33']
        sales_turnover_vill_agent2016_obj=worksheet['F33']
        sales_turnover_other2016_obj=worksheet['G33']
        sales_turnover_total2016_obj=worksheet['H33'] 
        
        num_unique_farmer2016_obj=worksheet['G34']
    
        worksheet = workbook['RETH-EN']
        print(f'Working on file: {file} ...')
        print('Sheet: RETH-EN ...')
        
        mgmt_capability_lev_obj=worksheet['D15']
        inst_competency_lev_obj=worksheet['D16']
        bus_knowledge_lev_obj=worksheet['D17']
        mark_potential_lev_obj=worksheet['D18']
        fin_management_lev_obj=worksheet['D19']
        env_compliance_lev_obj=worksheet['D20']
        mgmt_capability_lev_obj=worksheet['E15']
        inst_competency_lev_obj=worksheet['E16']
        bus_knowledge_lev_obj=worksheet['E17']
        mark_potential_lev_obj=worksheet['E18']
        fin_management_lev_obj=worksheet['E19']
        env_compliance_lev_obj=worksheet['E20']
      
        name_assessor.append(name_assessor_obj.value)
        date.append(date_obj.value)
        start_time.append(start_time_obj.value)
        end_time.append(end_time_obj.value)
        duration.append(duration_obj.value)
        type_assessed.append(type_assessed_obj.value)
        country.append(country_obj.value)
        code.append(code_obj.value)
        business_name.append(business_name_obj.value)
        legal_formation.append(legal_formation_obj.value)
        region.append(region_obj.value)
        zone.append(zone_obj.value)
        woreda.append(woreda_obj.value)
        telephone.append(telephone_obj.value)
        email.append(email_obj.value)
        contact_person.append(contact_person_obj.value)
        contact_designation.append(contact_designation_obj.value)
        contact_gender.append(contact_gender_obj.value)
        contact_educ.append(contact_educ_obj.value)
        hh_head_owner.append(hh_head_owner_obj.value)
        train_given_by.append(train_given_by_obj.value)
        train_satisfaction.append(train_satisfaction_obj.value)
        train_decision.append(train_decision_obj.value)
        have_license.append(have_license_obj.value)
        have_tin.append(have_tin_obj.value)
        year_estabilished.append(year_estabilished_obj.value)
        employment_male.append(employment_male_obj.value)
        employment_female.append(employment_female_obj.value)
        employment_total.append(employment_total_obj.value)
        avail_water.append(avail_water_obj.value)
        avail_input_store.append(avail_input_store_obj.value)
        avail_electricity.append(avail_electricity_obj.value)
        get_business_advice.append(get_business_advice_obj.value)
        access_poultry_mart.append(access_poultry_mart_obj.value)
        agency_access_loan.append(agency_access_loan_obj.value)
        agency_lender_type.append(agency_lender_type_obj.value)
        agency_amount.append(agency_amount_obj.value)
        informal_access_loan.append(informal_access_loan_obj.value)
        informal_lender_type.append(informal_lender_type_obj.value)
        informal_amount.append(informal_amount_obj.value)
        finance_challenge1.append(finance_challenge1_obj.value)
        finance_challenge2.append(finance_challenge2_obj.value)
        finance_challenge3.append(finance_challenge3_obj.value)
        finance_challenge4.append(finance_challenge4_obj.value)
        finance_challenge_other.append(finance_challenge_other_obj.value)
        
        level_educ.append(level_educ_obj.value)
        experience_years.append(experience_years_obj.value)
        received_training.append(received_training_obj.value)
        employment_procedure.append(employment_procedure_obj.value)
        provide_advisory_serv.append(provide_advisory_serv_obj.value)
        know_feed_types.append(know_feed_types_obj.value)
        know_feed_handling.append(know_feed_handling_obj.value)
        feed_store_standard.append(feed_store_standard_obj.value)
        store_sanitation.append(store_sanitation_obj.value)
        quality_control.append(quality_control_obj.value)
        have_annual_plan.append(have_annual_plan_obj.value)
        have_additional_income.append(have_additional_income_obj.value)
        alternative_finance.append(alternative_finance_obj.value)
        market_competition.append(market_competition_obj.value)
        promotion.append(promotion_obj.value)
        risk_price.append(risk_price_obj.value)
        risk_ss_shortage.append(risk_ss_shortage_obj.value)
        working_capital.append(working_capital_obj.value)
        financial_records.append(financial_records_obj.value)
        risk_potential.append(risk_potential_obj.value)
        mitigation_method.append(mitigation_method_obj.value)
        risk_health.append(risk_health_obj.value)
        num_purchase_inputs2018.append(num_purchase_inputs2018_obj.value)
        num_purchase_inputs2017.append(num_purchase_inputs2017_obj.value)
        num_purchase_inputs2016.append(num_purchase_inputs2016_obj.value)
        inputs_sold_gk_farmer2018.append(inputs_sold_gk_farmer2018_obj.value)
        inputs_sold_gk_gov_agent2018.append(inputs_sold_gk_gov_agent2018_obj.value)
        inputs_sold_gk_poult_agent2018.append(inputs_sold_gk_poult_agent2018_obj.value)
        inputs_sold_gk_vill_agent2018.append(inputs_sold_gk_vill_agent2018_obj.value)
        inputs_sold_gk_other2018.append(inputs_sold_gk_other2018_obj.value)
        inputs_sold_gk_total2018.append(inputs_sold_gk_total2018_obj.value)
        sales_turnover_farmer2018.append(sales_turnover_farmer2018_obj.value)
        sales_turnover_gov_agent2018.append(sales_turnover_gov_agent2018_obj.value)
        sales_turnover_poult_agent2018.append(sales_turnover_poult_agent2018_obj.value)
        sales_turnover_vill_agent2018.append(sales_turnover_vill_agent2018_obj.value)
        sales_turnover_other2018.append(sales_turnover_other2018_obj.value)
        sales_turnover_total2018.append(sales_turnover_total2018_obj.value)
        inputs_sold_gk_farmer2017.append(inputs_sold_gk_farmer2017_obj.value)
        inputs_sold_gk_gov_agent2017.append(inputs_sold_gk_gov_agent2017_obj.value)
        inputs_sold_gk_poult_agent2017.append(inputs_sold_gk_poult_agent2017_obj.value)
        inputs_sold_gk_vill_agent2017.append(inputs_sold_gk_vill_agent2017_obj.value)
        inputs_sold_gk_other2017.append(inputs_sold_gk_other2017_obj.value)
        inputs_sold_gk_total2017.append(inputs_sold_gk_total2017_obj.value)
        sales_turnover_farmer2017.append(sales_turnover_farmer2017_obj.value)
        sales_turnover_gov_agent2017.append(sales_turnover_gov_agent2017_obj.value)
        sales_turnover_poult_agent2017.append(sales_turnover_poult_agent2017_obj.value)
        sales_turnover_vill_agent2017.append(sales_turnover_vill_agent2017_obj.value)
        sales_turnover_other2017.append(sales_turnover_other2017_obj.value)
        sales_turnover_total2017.append(sales_turnover_total2017_obj.value)
        inputs_sold_gk_farmer2016.append(inputs_sold_gk_farmer2016_obj.value)
        inputs_sold_gk_gov_agent2016.append(inputs_sold_gk_gov_agent2016_obj.value)
        inputs_sold_gk_poult_agent2016.append(inputs_sold_gk_poult_agent2016_obj.value)
        inputs_sold_gk_vill_agent2016.append(inputs_sold_gk_vill_agent2016_obj.value)
        inputs_sold_gk_other2016.append(inputs_sold_gk_other2016_obj.value)
        inputs_sold_gk_total2016.append(inputs_sold_gk_total2016_obj.value)
        sales_turnover_farmer2016.append(sales_turnover_farmer2016_obj.value)
        sales_turnover_gov_agent2016.append(sales_turnover_gov_agent2016_obj.value)
        sales_turnover_poult_agent2016.append(sales_turnover_poult_agent2016_obj.value)
        sales_turnover_vill_agent2016.append(sales_turnover_vill_agent2016_obj.value)
        sales_turnover_other2016.append(sales_turnover_other2016_obj.value)
        sales_turnover_total2016.append(sales_turnover_total2016_obj.value)
        num_unique_farmer2018.append(num_unique_farmer2018_obj.value)
        num_unique_farmer2017.append(num_unique_farmer2017_obj.value)
        num_unique_farmer2016.append(num_unique_farmer2016_obj.value)
        
        mgmt_capability_lev.append(mgmt_capability_lev_obj.value)
        inst_competency_lev.append(inst_competency_lev_obj.value)
        bus_knowledge_lev.append(bus_knowledge_lev_obj.value)
        mark_potential_lev.append(mark_potential_lev_obj.value)
        fin_management_lev.append(fin_management_lev_obj.value)
        env_compliance_lev.append(env_compliance_lev_obj.value)
        mgmt_capability_lev.append(mgmt_capability_lev_obj.value)
        inst_competency_lev.append(inst_competency_lev_obj.value)
        bus_knowledge_lev.append(bus_knowledge_lev_obj.value)
        mark_potential_lev.append(mark_potential_lev_obj.value)
        fin_management_lev.append(fin_management_lev_obj.value)
        env_compliance_lev.append(env_compliance_lev_obj.value)
       
    data = {}
    
    data["pa_id"]=pa_id
    
    data["name_assessor"]=name_assessor
    data["date"]=date 
    data["start_time "]=start_time  
    data["end_time"]=end_time
    data["duration"]=duration
    data["type_assessed"]=type_assessed
    data["country"]=country
    data["code"]=code
    data["business_name"]=business_name
    data["legal_formation"]=legal_formation
    data["region"]=region
    data["zone"]=zone
    data["woreda"]=woreda
    data["telephone"]=telephone
    data["email"]=email
    data["contact_person"]=contact_person
    data["contact_designation"]=contact_designation
    data["contact_gender"]=contact_gender
    data["contact_educ"]=contact_educ
    data["hh_head_owner"]=hh_head_owner
    data["train_given_by"]=train_given_by
    data["train_satisfaction"]=train_satisfaction
    data["train_decision"]=train_decision
    data["have_license"]=have_license
    data["have_tin"]=have_tin
    data["year_estabilished"]=year_estabilished
    data["employment_male"]=employment_male 
    data["employment_female"]=employment_female 
    data["employment_total"]=employment_total
    data["avail_water"]=avail_water 
    data["avail_input_store"]=avail_input_store 
    data["avail_electricity"]=avail_electricity  
    data["get_business_advice"]=get_business_advice 
    data["access_poultry_mart"]=access_poultry_mart 
    data["agency_access_loan"]=agency_access_loan 
    data["agency_lender_type"]=agency_lender_type
    data["agency_amount"]=agency_amount 
    data["informal_access_loan"]=informal_access_loan 
    data["informal_lender_type"]=informal_lender_type 
    data["informal_amount"]=informal_amount
    data["finance_challenge1"]=finance_challenge1 
    data["finance_challenge2"]=finance_challenge2 
    data["finance_challenge3"]=finance_challenge3
    data["finance_challenge4"]=finance_challenge4
    data["finance_challenge_other"]=finance_challenge_other
    
    data["level_educ"]=level_educ 
    data["experience_years"]=experience_years 
    data["received_training"]=received_training 
    data["employment_procedure"]=employment_procedure 
    data["provide_advisory_serv"]=provide_advisory_serv 
    data["know_feed_types"]=know_feed_types
    data["know_feed_handling"]=know_feed_handling
    data["feed_store_standard"]=feed_store_standard
    data["store_sanitation"]=store_sanitation
    data["quality_control"]=quality_control
    data["have_annual_plan"]=have_annual_plan 
    data["have_additional_income"]=have_additional_income  
    data["alternative_finance"]=alternative_finance 
    data["market_competition"]=market_competition 
    data["promotion"]=promotion 
    data["risk_price"]=risk_price 
    data["risk_ss_shortage"]=risk_ss_shortage
    data["working_capital"]=working_capital 
    data["financial_records"]=financial_records 
    data["risk_potential"]=risk_potential
    data["mitigation_method"]=mitigation_method 
    data["risk_health"]=risk_health
    
    data["num_purchase_inputs2018"]=num_purchase_inputs2018 
    data["num_purchase_inputs2017"]=num_purchase_inputs2017 
    data["num_purchase_inputs2016"]=num_purchase_inputs2016
    data["inputs_sold_gk_farmer2018"]=inputs_sold_gk_farmer2018
    data["inputs_sold_gk_gov_agent2018"]=inputs_sold_gk_gov_agent2018
    data["inputs_sold_gk_poult_agent2018"]=inputs_sold_gk_poult_agent2018
    data["inputs_sold_gk_vill_agent2018"]=inputs_sold_gk_vill_agent2018
    data["inputs_sold_gk_other2018"]=inputs_sold_gk_other2018
    data["inputs_sold_gk_total2018"]=inputs_sold_gk_total2018
    data["sales_turnover_farmer2018"]=sales_turnover_farmer2018
    data["sales_turnover_gov_agent2018"]=sales_turnover_gov_agent2018
    data["sales_turnover_poult_agent2018"]=sales_turnover_poult_agent2018
    data["sales_turnover_vill_agent2018"]=sales_turnover_vill_agent2018
    data["sales_turnover_other2018"]=sales_turnover_other2018
    data["sales_turnover_total2018"]=sales_turnover_total2018
    data["inputs_sold_gk_farmer2017"]=inputs_sold_gk_farmer2017
    data["inputs_sold_gk_gov_agent2017"]=inputs_sold_gk_gov_agent2017
    data["inputs_sold_gk_poult_agent2017"]=inputs_sold_gk_poult_agent2017
    data["inputs_sold_gk_vill_agent2017"]=inputs_sold_gk_vill_agent2017
    data["inputs_sold_gk_other2017"]=inputs_sold_gk_other2017
    data["inputs_sold_gk_total2017"]=inputs_sold_gk_total2017
    data["sales_turnover_farmer2017"]=sales_turnover_farmer2017
    data["sales_turnover_gov_agent2017"]=sales_turnover_gov_agent2017
    data["sales_turnover_poult_agent2017"]=sales_turnover_poult_agent2017
    data["sales_turnover_vill_agent2017"]=sales_turnover_vill_agent2017
    data["sales_turnover_other2017"]=sales_turnover_other2017
    data["sales_turnover_total2017"]=sales_turnover_total2017
    data["inputs_sold_gk_farmer2016"]=inputs_sold_gk_farmer2016
    data["inputs_sold_gk_gov_agent2016"]=inputs_sold_gk_gov_agent2016
    data["inputs_sold_gk_poult_agent2016"]=inputs_sold_gk_poult_agent2016
    data["inputs_sold_gk_vill_agent2016"]=inputs_sold_gk_vill_agent2016
    data["inputs_sold_gk_other2016"]=inputs_sold_gk_other2016
    data["inputs_sold_gk_total2016"]=inputs_sold_gk_total2016
    data["sales_turnover_farmer2016"]=sales_turnover_farmer2016
    data["sales_turnover_gov_agent2016"]=sales_turnover_gov_agent2016
    data["sales_turnover_poult_agent2016"]=sales_turnover_poult_agent2016
    data["sales_turnover_vill_agent2016"]=sales_turnover_vill_agent2016
    data["sales_turnover_other2016"]=sales_turnover_other2016
    data["sales_turnover_total2016"]=sales_turnover_total2016
    data["num_unique_farmer2018"]=num_unique_farmer2018
    data["num_unique_farmer2017"]=num_unique_farmer2017
    data["num_unique_farmer2016"]=num_unique_farmer2016
    
    data["mgmt_capability_lev"]=mgmt_capability_lev
    data["inst_competency_lev"]=inst_competency_lev
    data["bus_knowledge_lev"]=bus_knowledge_lev
    data["mark_potential_lev"]=mark_potential_lev
    data["fin_management_lev"]=fin_management_lev
    data["env_compliance_lev"]=env_compliance_lev
    data["mgmt_capability_lev"]=mgmt_capability_lev
    data["inst_competency_lev"]=inst_competency_lev
    data["bus_knowledge_lev"]=bus_knowledge_lev
    data["mark_potential_lev"]=mark_potential_lev
    data["fin_management_lev"]=fin_management_lev
    data["env_compliance_lev"]=env_compliance_lev
    
    df = pd.DataFrame.from_dict(data) 
    return df




def get_data_fd2(file_path, files):
    [
    pa_id,
    name_assessor,
    date ,
    start_time,  
    end_time,
    duration,
    type_assessed,
    country,
    code,
    business_name,
    legal_formation,
    region,
    zone,
    woreda,
    telephone,
    email,
    contact_person,
    contact_designation,
    contact_gender,
    contact_educ,
    hh_head_owner,
    train_given_by,
    train_satisfaction,
    train_decision,
    have_license,
    have_tin,
    year_estabilished,
    employment_male, 
    employment_female, 
    employment_total,
    avail_water, 
    avail_input_store, 
    avail_electricity,  
    get_business_advice, 
    access_poultry_mart, 
    agency_access_loan, 
    agency_lender_type,
    agency_amount, 
    informal_access_loan, 
    informal_lender_type, 
    informal_amount,
    finance_challenge1, 
    finance_challenge2, 
    finance_challenge3,
    finance_challenge4,
    finance_challenge_other, 
    level_educ, 
    experience_years, 
    received_training, 
    employment_procedure, 
    provide_advisory_serv, 
    know_feed_types,
    know_feed_handling,
    feed_store_standard,
    store_sanitation,
    quality_control,
    have_annual_plan, 
    have_additional_income,  
    alternative_finance, 
    market_competition, 
    promotion, 
    risk_price, 
    risk_ss_shortage,
    working_capital, 
    financial_records, 
    risk_potential,
    mitigation_method, 
    risk_health,
    num_purchase_inputs2018, 
    num_purchase_inputs2017, 
    num_purchase_inputs2016,
    inputs_sold_gk_farmer2018,
    inputs_sold_gk_gov_agent2018,
    inputs_sold_gk_poult_agent2018,
    inputs_sold_gk_vill_agent2018,
    inputs_sold_gk_other2018,
    inputs_sold_gk_total2018,
    sales_turnover_farmer2018,
    sales_turnover_gov_agent2018,
    sales_turnover_poult_agent2018,
    sales_turnover_vill_agent2018,
    sales_turnover_other2018,
    sales_turnover_total2018,
    inputs_sold_gk_farmer2017,
    inputs_sold_gk_gov_agent2017,
    inputs_sold_gk_poult_agent2017,
    inputs_sold_gk_vill_agent2017,
    inputs_sold_gk_other2017,
    inputs_sold_gk_total2017,
    sales_turnover_farmer2017,
    sales_turnover_gov_agent2017,
    sales_turnover_poult_agent2017,
    sales_turnover_vill_agent2017,
    sales_turnover_other2017,
    sales_turnover_total2017,
    inputs_sold_gk_farmer2016,
    inputs_sold_gk_gov_agent2016,
    inputs_sold_gk_poult_agent2016,
    inputs_sold_gk_vill_agent2016,
    inputs_sold_gk_other2016,
    inputs_sold_gk_total2016,
    sales_turnover_farmer2016,
    sales_turnover_gov_agent2016,
    sales_turnover_poult_agent2016,
    sales_turnover_vill_agent2016,
    sales_turnover_other2016,
    sales_turnover_total2016,
    num_unique_farmer2018,
    num_unique_farmer2017,
    num_unique_farmer2016,
    
    mgmt_capability_lev,
    inst_competency_lev,
    bus_knowledge_lev,
    mark_potential_lev,
    fin_management_lev,
    env_compliance_lev,
    mgmt_capability_lev,
    inst_competency_lev,
    bus_knowledge_lev,
    mark_potential_lev,
    fin_management_lev,
    env_compliance_lev
    ]= ([] for i in range(122))
    
    
    for file in files:
        match = re.search(r'\d*', file)
        pa_id.append(match.group())
        
        workbook = load_workbook(file_path+file, data_only=True)
        
        worksheet = workbook['BIETH']
        print(f'Working on file: {file} ...')
        print('Sheet: BIETH ...')
         
        name_assessor_obj=worksheet['G6']
        date_obj=worksheet['A10']
        start_time_obj=worksheet['B10']
        end_time_obj=worksheet['C10']
        duration_obj=worksheet['D10']
        
        type_assessed_obj=worksheet['F9']
        country_obj=worksheet['G9']
        code_obj=worksheet['H9']
        
        business_name_obj=worksheet['C15']
        legal_formation_obj=worksheet['C16']
        region_obj=worksheet['C17']
        zone_obj=worksheet['C18']
        woreda_obj=worksheet['C19']
        telephone_obj=worksheet['C20']
        email_obj=worksheet['C21']
        
        contact_person_obj=worksheet['G15']
        contact_designation_obj=worksheet['H15']
        contact_gender_obj=worksheet['G16']
        contact_educ_obj=worksheet['G17']
        hh_head_owner_obj=worksheet['G19']
        
        train_given_by_obj=worksheet['C24']
        train_satisfaction_obj=worksheet['F24']
        train_decision_obj=worksheet['G24']
        
        have_license_obj=worksheet['C31']
        have_tin_obj=worksheet['C32']
        year_estabilished_obj=worksheet['G31']
        
        employment_male_obj=worksheet['A36']
        employment_female_obj=worksheet['B36']
        employment_total_obj=worksheet['C36']
        avail_water_obj=worksheet['H34']
        avail_input_store_obj=worksheet['H35']
        avail_electricity_obj=worksheet['H36']
        
        get_business_advice_obj=worksheet['A40']
        access_poultry_mart_obj=worksheet['F40']
        
        agency_access_loan_obj=worksheet['C42']
        agency_lender_type_obj=worksheet['F42']
        agency_amount_obj=worksheet['H42']
        informal_access_loan_obj=worksheet['C43']
        informal_lender_type_obj=worksheet['F43']
        informal_amount_obj=worksheet['H43']
        
        finance_challenge1_obj=worksheet['A46']
        finance_challenge2_obj=worksheet['C46']
        finance_challenge3_obj=worksheet['E46']
        finance_challenge4_obj=worksheet['G46']
        finance_challenge_other_obj=worksheet['C47']
        
        worksheet = workbook['QETH']
        print(f'Working on file: {file} ...')
        print('Sheet: QETH ...')
    
        level_educ_obj=worksheet['F14']
        experience_years_obj=worksheet['F20']
        received_training_obj=worksheet['F26']
        employment_procedure_obj=worksheet['F33']
        provide_advisory_serv_obj=worksheet['F39']
        
        know_feed_types_obj=worksheet['F49']
        know_feed_handling_obj=worksheet['F55']
        feed_store_standard_obj=worksheet['F62']
        store_sanitation_obj=worksheet['F68']
        quality_control_obj=worksheet['F75']
        have_annual_plan_obj=worksheet['F85']
        have_additional_income_obj=worksheet['F92']
        alternative_finance_obj=worksheet['F99']
        market_competition_obj=worksheet['F109']
        promotion_obj=worksheet['F115']
        risk_price_obj=worksheet['F122']
        risk_ss_shortage_obj=worksheet['F128']
        working_capital_obj=worksheet['F138']
        financial_records_obj=worksheet['F144']
        risk_potential_obj=worksheet['F154']
        mitigation_method_obj=worksheet['F160']
        risk_health_obj=worksheet['F167']
        
        worksheet = workbook['Additional Info']
        print(f'Working on file: {file} ...')
        print('Sheet: Additional Info ...')
        
        num_purchase_inputs2018_obj=worksheet['C15']
        num_purchase_inputs2017_obj=worksheet['D15']
        num_purchase_inputs2016_obj=worksheet['E15']    
        
        inputs_sold_gk_farmer2018_obj=worksheet['C20']
        inputs_sold_gk_gov_agent2018_obj=worksheet['D20']
        inputs_sold_gk_poult_agent2018_obj=worksheet['E20']
        inputs_sold_gk_vill_agent2018_obj=worksheet['F20']
        inputs_sold_gk_other2018_obj=worksheet['G20']
        inputs_sold_gk_total2018_obj=worksheet['H20']    
        
        sales_turnover_farmer2018_obj=worksheet['C21']
        sales_turnover_gov_agent2018_obj=worksheet['D21']
        sales_turnover_poult_agent2018_obj=worksheet['E21']
        sales_turnover_vill_agent2018_obj=worksheet['F21']
        sales_turnover_other2018_obj=worksheet['G21']
        sales_turnover_total2018_obj=worksheet['H21']    
    
        num_unique_farmer2018_obj=worksheet['G22']
        
        inputs_sold_gk_farmer2017_obj=worksheet['C26']
        inputs_sold_gk_gov_agent2017_obj=worksheet['D26']
        inputs_sold_gk_poult_agent2017_obj=worksheet['E26']
        inputs_sold_gk_vill_agent2017_obj=worksheet['F26']
        inputs_sold_gk_other2017_obj=worksheet['G26']
        inputs_sold_gk_total2017_obj=worksheet['H26']    
        
        sales_turnover_farmer2017_obj=worksheet['C27']
        sales_turnover_gov_agent2017_obj=worksheet['D27']
        sales_turnover_poult_agent2017_obj=worksheet['E27']
        sales_turnover_vill_agent2017_obj=worksheet['F27']
        sales_turnover_other2017_obj=worksheet['G27']
        sales_turnover_total2017_obj=worksheet['H27']    
    
        num_unique_farmer2017_obj=worksheet['G28']
        
        inputs_sold_gk_farmer2016_obj=worksheet['C32']
        inputs_sold_gk_gov_agent2016_obj=worksheet['D32']
        inputs_sold_gk_poult_agent2016_obj=worksheet['E32']
        inputs_sold_gk_vill_agent2016_obj=worksheet['F32']
        inputs_sold_gk_other2016_obj=worksheet['G32']
        inputs_sold_gk_total2016_obj=worksheet['H32'] 
        
        sales_turnover_farmer2016_obj=worksheet['C33']
        sales_turnover_gov_agent2016_obj=worksheet['D33']
        sales_turnover_poult_agent2016_obj=worksheet['E33']
        sales_turnover_vill_agent2016_obj=worksheet['F33']
        sales_turnover_other2016_obj=worksheet['G33']
        sales_turnover_total2016_obj=worksheet['H33'] 
        
        num_unique_farmer2016_obj=worksheet['G34']
    
        worksheet = workbook['RETH-EN']
        print(f'Working on file: {file} ...')
        print('Sheet: RETH-EN ...')
        
        mgmt_capability_lev_obj=worksheet['D15']
        inst_competency_lev_obj=worksheet['D16']
        bus_knowledge_lev_obj=worksheet['D17']
        mark_potential_lev_obj=worksheet['D18']
        fin_management_lev_obj=worksheet['D19']
        env_compliance_lev_obj=worksheet['D20']
        mgmt_capability_lev_obj=worksheet['E15']
        inst_competency_lev_obj=worksheet['E16']
        bus_knowledge_lev_obj=worksheet['E17']
        mark_potential_lev_obj=worksheet['E18']
        fin_management_lev_obj=worksheet['E19']
        env_compliance_lev_obj=worksheet['E20']
      
        name_assessor.append(name_assessor_obj.value)
        date.append(date_obj.value)
        start_time.append(start_time_obj.value)
        end_time.append(end_time_obj.value)
        duration.append(duration_obj.value)
        type_assessed.append(type_assessed_obj.value)
        country.append(country_obj.value)
        code.append(code_obj.value)
        business_name.append(business_name_obj.value)
        legal_formation.append(legal_formation_obj.value)
        region.append(region_obj.value)
        zone.append(zone_obj.value)
        woreda.append(woreda_obj.value)
        telephone.append(telephone_obj.value)
        email.append(email_obj.value)
        contact_person.append(contact_person_obj.value)
        contact_designation.append(contact_designation_obj.value)
        contact_gender.append(contact_gender_obj.value)
        contact_educ.append(contact_educ_obj.value)
        hh_head_owner.append(hh_head_owner_obj.value)
        train_given_by.append(train_given_by_obj.value)
        train_satisfaction.append(train_satisfaction_obj.value)
        train_decision.append(train_decision_obj.value)
        have_license.append(have_license_obj.value)
        have_tin.append(have_tin_obj.value)
        year_estabilished.append(year_estabilished_obj.value)
        employment_male.append(employment_male_obj.value)
        employment_female.append(employment_female_obj.value)
        employment_total.append(employment_total_obj.value)
        avail_water.append(avail_water_obj.value)
        avail_input_store.append(avail_input_store_obj.value)
        avail_electricity.append(avail_electricity_obj.value)
        get_business_advice.append(get_business_advice_obj.value)
        access_poultry_mart.append(access_poultry_mart_obj.value)
        agency_access_loan.append(agency_access_loan_obj.value)
        agency_lender_type.append(agency_lender_type_obj.value)
        agency_amount.append(agency_amount_obj.value)
        informal_access_loan.append(informal_access_loan_obj.value)
        informal_lender_type.append(informal_lender_type_obj.value)
        informal_amount.append(informal_amount_obj.value)
        finance_challenge1.append(finance_challenge1_obj.value)
        finance_challenge2.append(finance_challenge2_obj.value)
        finance_challenge3.append(finance_challenge3_obj.value)
        finance_challenge4.append(finance_challenge4_obj.value)
        finance_challenge_other.append(finance_challenge_other_obj.value)
        
        level_educ.append(level_educ_obj.value)
        experience_years.append(experience_years_obj.value)
        received_training.append(received_training_obj.value)
        employment_procedure.append(employment_procedure_obj.value)
        provide_advisory_serv.append(provide_advisory_serv_obj.value)
        know_feed_types.append(know_feed_types_obj.value)
        know_feed_handling.append(know_feed_handling_obj.value)
        feed_store_standard.append(feed_store_standard_obj.value)
        store_sanitation.append(store_sanitation_obj.value)
        quality_control.append(quality_control_obj.value)
        have_annual_plan.append(have_annual_plan_obj.value)
        have_additional_income.append(have_additional_income_obj.value)
        alternative_finance.append(alternative_finance_obj.value)
        market_competition.append(market_competition_obj.value)
        promotion.append(promotion_obj.value)
        risk_price.append(risk_price_obj.value)
        risk_ss_shortage.append(risk_ss_shortage_obj.value)
        working_capital.append(working_capital_obj.value)
        financial_records.append(financial_records_obj.value)
        risk_potential.append(risk_potential_obj.value)
        mitigation_method.append(mitigation_method_obj.value)
        risk_health.append(risk_health_obj.value)
        num_purchase_inputs2018.append(num_purchase_inputs2018_obj.value)
        num_purchase_inputs2017.append(num_purchase_inputs2017_obj.value)
        num_purchase_inputs2016.append(num_purchase_inputs2016_obj.value)
        inputs_sold_gk_farmer2018.append(inputs_sold_gk_farmer2018_obj.value)
        inputs_sold_gk_gov_agent2018.append(inputs_sold_gk_gov_agent2018_obj.value)
        inputs_sold_gk_poult_agent2018.append(inputs_sold_gk_poult_agent2018_obj.value)
        inputs_sold_gk_vill_agent2018.append(inputs_sold_gk_vill_agent2018_obj.value)
        inputs_sold_gk_other2018.append(inputs_sold_gk_other2018_obj.value)
        inputs_sold_gk_total2018.append(inputs_sold_gk_total2018_obj.value)
        sales_turnover_farmer2018.append(sales_turnover_farmer2018_obj.value)
        sales_turnover_gov_agent2018.append(sales_turnover_gov_agent2018_obj.value)
        sales_turnover_poult_agent2018.append(sales_turnover_poult_agent2018_obj.value)
        sales_turnover_vill_agent2018.append(sales_turnover_vill_agent2018_obj.value)
        sales_turnover_other2018.append(sales_turnover_other2018_obj.value)
        sales_turnover_total2018.append(sales_turnover_total2018_obj.value)
        inputs_sold_gk_farmer2017.append(inputs_sold_gk_farmer2017_obj.value)
        inputs_sold_gk_gov_agent2017.append(inputs_sold_gk_gov_agent2017_obj.value)
        inputs_sold_gk_poult_agent2017.append(inputs_sold_gk_poult_agent2017_obj.value)
        inputs_sold_gk_vill_agent2017.append(inputs_sold_gk_vill_agent2017_obj.value)
        inputs_sold_gk_other2017.append(inputs_sold_gk_other2017_obj.value)
        inputs_sold_gk_total2017.append(inputs_sold_gk_total2017_obj.value)
        sales_turnover_farmer2017.append(sales_turnover_farmer2017_obj.value)
        sales_turnover_gov_agent2017.append(sales_turnover_gov_agent2017_obj.value)
        sales_turnover_poult_agent2017.append(sales_turnover_poult_agent2017_obj.value)
        sales_turnover_vill_agent2017.append(sales_turnover_vill_agent2017_obj.value)
        sales_turnover_other2017.append(sales_turnover_other2017_obj.value)
        sales_turnover_total2017.append(sales_turnover_total2017_obj.value)
        inputs_sold_gk_farmer2016.append(inputs_sold_gk_farmer2016_obj.value)
        inputs_sold_gk_gov_agent2016.append(inputs_sold_gk_gov_agent2016_obj.value)
        inputs_sold_gk_poult_agent2016.append(inputs_sold_gk_poult_agent2016_obj.value)
        inputs_sold_gk_vill_agent2016.append(inputs_sold_gk_vill_agent2016_obj.value)
        inputs_sold_gk_other2016.append(inputs_sold_gk_other2016_obj.value)
        inputs_sold_gk_total2016.append(inputs_sold_gk_total2016_obj.value)
        sales_turnover_farmer2016.append(sales_turnover_farmer2016_obj.value)
        sales_turnover_gov_agent2016.append(sales_turnover_gov_agent2016_obj.value)
        sales_turnover_poult_agent2016.append(sales_turnover_poult_agent2016_obj.value)
        sales_turnover_vill_agent2016.append(sales_turnover_vill_agent2016_obj.value)
        sales_turnover_other2016.append(sales_turnover_other2016_obj.value)
        sales_turnover_total2016.append(sales_turnover_total2016_obj.value)
        num_unique_farmer2018.append(num_unique_farmer2018_obj.value)
        num_unique_farmer2017.append(num_unique_farmer2017_obj.value)
        num_unique_farmer2016.append(num_unique_farmer2016_obj.value)
        
        mgmt_capability_lev.append(mgmt_capability_lev_obj.value)
        inst_competency_lev.append(inst_competency_lev_obj.value)
        bus_knowledge_lev.append(bus_knowledge_lev_obj.value)
        mark_potential_lev.append(mark_potential_lev_obj.value)
        fin_management_lev.append(fin_management_lev_obj.value)
        env_compliance_lev.append(env_compliance_lev_obj.value)
        mgmt_capability_lev.append(mgmt_capability_lev_obj.value)
        inst_competency_lev.append(inst_competency_lev_obj.value)
        bus_knowledge_lev.append(bus_knowledge_lev_obj.value)
        mark_potential_lev.append(mark_potential_lev_obj.value)
        fin_management_lev.append(fin_management_lev_obj.value)
        env_compliance_lev.append(env_compliance_lev_obj.value)
       
    data = {}
    
    data["pa_id"]=pa_id
    
    data["name_assessor"]=name_assessor
    data["date"]=date 
    data["start_time "]=start_time  
    data["end_time"]=end_time
    data["duration"]=duration
    data["type_assessed"]=type_assessed
    data["country"]=country
    data["code"]=code
    data["business_name"]=business_name
    data["legal_formation"]=legal_formation
    data["region"]=region
    data["zone"]=zone
    data["woreda"]=woreda
    data["telephone"]=telephone
    data["email"]=email
    data["contact_person"]=contact_person
    data["contact_designation"]=contact_designation
    data["contact_gender"]=contact_gender
    data["contact_educ"]=contact_educ
    data["hh_head_owner"]=hh_head_owner
    data["train_given_by"]=train_given_by
    data["train_satisfaction"]=train_satisfaction
    data["train_decision"]=train_decision
    data["have_license"]=have_license
    data["have_tin"]=have_tin
    data["year_estabilished"]=year_estabilished
    data["employment_male"]=employment_male 
    data["employment_female"]=employment_female 
    data["employment_total"]=employment_total
    data["avail_water"]=avail_water 
    data["avail_input_store"]=avail_input_store 
    data["avail_electricity"]=avail_electricity  
    data["get_business_advice"]=get_business_advice 
    data["access_poultry_mart"]=access_poultry_mart 
    data["agency_access_loan"]=agency_access_loan 
    data["agency_lender_type"]=agency_lender_type
    data["agency_amount"]=agency_amount 
    data["informal_access_loan"]=informal_access_loan 
    data["informal_lender_type"]=informal_lender_type 
    data["informal_amount"]=informal_amount
    data["finance_challenge1"]=finance_challenge1 
    data["finance_challenge2"]=finance_challenge2 
    data["finance_challenge3"]=finance_challenge3
    data["finance_challenge4"]=finance_challenge4
    data["finance_challenge_other"]=finance_challenge_other
    
    data["level_educ"]=level_educ 
    data["experience_years"]=experience_years 
    data["received_training"]=received_training 
    data["employment_procedure"]=employment_procedure 
    data["provide_advisory_serv"]=provide_advisory_serv 
    data["know_feed_types"]=know_feed_types
    data["know_feed_handling"]=know_feed_handling
    data["feed_store_standard"]=feed_store_standard
    data["store_sanitation"]=store_sanitation
    data["quality_control"]=quality_control
    data["have_annual_plan"]=have_annual_plan 
    data["have_additional_income"]=have_additional_income  
    data["alternative_finance"]=alternative_finance 
    data["market_competition"]=market_competition 
    data["promotion"]=promotion 
    data["risk_price"]=risk_price 
    data["risk_ss_shortage"]=risk_ss_shortage
    data["working_capital"]=working_capital 
    data["financial_records"]=financial_records 
    data["risk_potential"]=risk_potential
    data["mitigation_method"]=mitigation_method 
    data["risk_health"]=risk_health
    
    data["num_purchase_inputs2018"]=num_purchase_inputs2018 
    data["num_purchase_inputs2017"]=num_purchase_inputs2017 
    data["num_purchase_inputs2016"]=num_purchase_inputs2016
    data["inputs_sold_gk_farmer2018"]=inputs_sold_gk_farmer2018
    data["inputs_sold_gk_gov_agent2018"]=inputs_sold_gk_gov_agent2018
    data["inputs_sold_gk_poult_agent2018"]=inputs_sold_gk_poult_agent2018
    data["inputs_sold_gk_vill_agent2018"]=inputs_sold_gk_vill_agent2018
    data["inputs_sold_gk_other2018"]=inputs_sold_gk_other2018
    data["inputs_sold_gk_total2018"]=inputs_sold_gk_total2018
    data["sales_turnover_farmer2018"]=sales_turnover_farmer2018
    data["sales_turnover_gov_agent2018"]=sales_turnover_gov_agent2018
    data["sales_turnover_poult_agent2018"]=sales_turnover_poult_agent2018
    data["sales_turnover_vill_agent2018"]=sales_turnover_vill_agent2018
    data["sales_turnover_other2018"]=sales_turnover_other2018
    data["sales_turnover_total2018"]=sales_turnover_total2018
    data["inputs_sold_gk_farmer2017"]=inputs_sold_gk_farmer2017
    data["inputs_sold_gk_gov_agent2017"]=inputs_sold_gk_gov_agent2017
    data["inputs_sold_gk_poult_agent2017"]=inputs_sold_gk_poult_agent2017
    data["inputs_sold_gk_vill_agent2017"]=inputs_sold_gk_vill_agent2017
    data["inputs_sold_gk_other2017"]=inputs_sold_gk_other2017
    data["inputs_sold_gk_total2017"]=inputs_sold_gk_total2017
    data["sales_turnover_farmer2017"]=sales_turnover_farmer2017
    data["sales_turnover_gov_agent2017"]=sales_turnover_gov_agent2017
    data["sales_turnover_poult_agent2017"]=sales_turnover_poult_agent2017
    data["sales_turnover_vill_agent2017"]=sales_turnover_vill_agent2017
    data["sales_turnover_other2017"]=sales_turnover_other2017
    data["sales_turnover_total2017"]=sales_turnover_total2017
    data["inputs_sold_gk_farmer2016"]=inputs_sold_gk_farmer2016
    data["inputs_sold_gk_gov_agent2016"]=inputs_sold_gk_gov_agent2016
    data["inputs_sold_gk_poult_agent2016"]=inputs_sold_gk_poult_agent2016
    data["inputs_sold_gk_vill_agent2016"]=inputs_sold_gk_vill_agent2016
    data["inputs_sold_gk_other2016"]=inputs_sold_gk_other2016
    data["inputs_sold_gk_total2016"]=inputs_sold_gk_total2016
    data["sales_turnover_farmer2016"]=sales_turnover_farmer2016
    data["sales_turnover_gov_agent2016"]=sales_turnover_gov_agent2016
    data["sales_turnover_poult_agent2016"]=sales_turnover_poult_agent2016
    data["sales_turnover_vill_agent2016"]=sales_turnover_vill_agent2016
    data["sales_turnover_other2016"]=sales_turnover_other2016
    data["sales_turnover_total2016"]=sales_turnover_total2016
    data["num_unique_farmer2018"]=num_unique_farmer2018
    data["num_unique_farmer2017"]=num_unique_farmer2017
    data["num_unique_farmer2016"]=num_unique_farmer2016
    
    data["mgmt_capability_lev"]=mgmt_capability_lev
    data["inst_competency_lev"]=inst_competency_lev
    data["bus_knowledge_lev"]=bus_knowledge_lev
    data["mark_potential_lev"]=mark_potential_lev
    data["fin_management_lev"]=fin_management_lev
    data["env_compliance_lev"]=env_compliance_lev
    data["mgmt_capability_lev"]=mgmt_capability_lev
    data["inst_competency_lev"]=inst_competency_lev
    data["bus_knowledge_lev"]=bus_knowledge_lev
    data["mark_potential_lev"]=mark_potential_lev
    data["fin_management_lev"]=fin_management_lev
    data["env_compliance_lev"]=env_compliance_lev
  
    df = pd.DataFrame.from_dict(data) 
    return df