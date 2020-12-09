# -*- coding: utf-8 -*-
"""
Created on Sat Nov 14 18:22:02 2020

@author: Admin
"""


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter


data="D:/projects/IFC/second_round/data/modified/" 
tables = "D:/projects/IFC/second_round/reports/alemu/tables"


df=pd.read_stata(data+'assessment_result_PAs_all_long.dta')


cont_vars= ['employment_male', 'employment_female', 'employment_total',
 'agency_amount', 'informal_amount', 'mgmt_capability_score',
 'inst_competency_score', 'bus_knowledge_score', 'mark_potential_score',
 'fin_management_score', 'env_compliance_score', 'average_score', 
 'prod_cycle', 'num_chicks_sold_farm', 'num_chicks_sold_gov_agent', 
 'num_chicks_sold_trader', 'num_chicks_sold_vill_agents',
 'num_chicks_sold_other', 'num_chicks_sold_total',
 'turn_over_farm', 'turn_over_gov_agent', 'turn_over_trader',
 'turn_over_vill_agents', 'turn_over_other', 'turn_over_total',
 'sold_to_unique_farmers']


cat_vars=['education', 'experience', 'training', 'staffing', 'customer_service',
  'housing', 'poultry_prtection_inputs', 'feed_and_feeding',
  'water_source_use', 'chicken_breed_quality', 'poultry_pests',
  'farm_hygen', 'annual_plan', 'income_source', 'financial_reliability',
  'competition', 'promotion', 'price_risk', 'supply_shortage',
  'working_capital', 'financial_records', 'environmental_threats',
  'mitigation_emvironmental', 'health_and_safety', 'contact_gender',
  'have_license', 'have_tin', 'contact_designation', 'legal_formation',
  'avail_water', 'avail_input_store', 'avail_electricity',
  'avail_generator', 'avail_heating', 'agency_access_loan',
  'agency_lender_type', 'informal_access_loan',
  'informal_lender_type', 'contact_educ_cat',
  'train_given_by', 'train_satisfaction', 'train_decision',
  'water_supply_man', 'water_supply_woman', 'water_supply_children',
  'water_supply_all', 'source_prot_inputs', 'get_business_advice',
  'access_poultry_mart', 'finance_challenge1', 'finance_challenge2',
  'finance_challenge3', 'finance_challenge4']

###############################################################################






################# PRODUCING TABLES #########################################
pattern = PatternFill(fill_type='solid',  start_color='00FF9900')
align = Alignment(horizontal='left', vertical='center')



###################### BEGIN: SUMMARY STATISTICS ##############################

writer=pd.ExcelWriter(tables+'summary_statistics.xlsx', engine='openpyxl')


#Basic summary statistics for continuous variables
df[cont_vars].describe().round(2).to_excel(writer, sheet_name= 'summary_stats')



#Average for continuous variables over region
region=pd.pivot_table(df, index=['region']).round(2)
region.drop(['year', 'round', 'code'], axis='columns', inplace=True)
region.to_excel(writer, sheet_name='average_region')

region_year=pd.pivot_table(df, index=['region', 'year']).round(2)
region_year.drop(['round', 'code', ], axis='columns', inplace=True)
region_year.to_excel(writer, sheet_name='average_region_year')


#Average for continuous variables over region and level of education
region_education=pd.pivot_table(df, index=['region', 'education']).round(2)
region_education.drop(['year', 'round', 'code'], axis='columns', inplace=True)
region_education.to_excel(writer, sheet_name='average_region_education')



#Pearson Correlation coefficient for continuous variables 
df[cont_vars].corr().round(2).to_excel(writer, sheet_name='correlation')

writer.save()
writer.close()
########################### END: SUMMARY STATISTICS ############################









######### BEGIN: Formmating tables ###############################3
book = load_workbook(tables+'summary_statistics.xlsx')
from openpyxl.formatting.rule import CellIsRule

color1 = PatternFill(start_color='d7c797',
                      end_color='d7c797',
                      fill_type='solid')

color2 = PatternFill(start_color='845422',
                      end_color='845422',
                      fill_type='solid')


for sheet in book.sheetnames:
    sheet = book[sheet]
    
    for column_cells in sheet.columns:
        length = max(map(lambda cell: len(str(cell.value)) if cell.value else 0, column_cells))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width=length+2
        
    for r in sheet.rows:
        cells= r[:1][0]
        cells.alignment=align
        
    for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.fill=pattern
    
    average_region_education=book['average_region_education']
    for r in average_region_education.rows:
        cells= r[1:2][0]
        cells.alignment=align
        
correlation = book['correlation']
correlation['A1']='Variables'

correlation['A29']= "Keys"
correlation['A30']='Color'
correlation['B30']='Correlation'
correlation['A31'].fill=color1
correlation['A32'].fill=color2
correlation['B31']= "Positively Correlated"
correlation['B32']= "Nagatively correlated"


correlation.conditional_formatting.add('B2:AA27',
                                       CellIsRule(operator='between',
                                                  formula=['0.5', '1'],
                                                  stopIfTrue=True,
                                                  fill=color1))  

correlation.conditional_formatting.add('B2:AA27',
                                       CellIsRule(operator='between',
                                                  formula=['-0.5', '-1'],
                                                  stopIfTrue=True,
                                                  fill=color2))

book.save(tables+'summary_statistics.xlsx')
######### END: Formmating tables ##########################################







##############BEGIN: TABULATION  FOR CATEGORICAL VARIABLES#################
writer=pd.ExcelWriter(tables+'tabulation.xlsx', engine='openpyxl')

for c in cat_vars:
    res = df[c].value_counts().to_frame('Frequency')
    res['Percent (%)'] = res['Frequency'] / res['Frequency'].sum()
    res['Percent (%)'] = res['Percent (%)'].mul(100).round(2)
    res.to_excel(writer, sheet_name=c)
    writer.save()    
writer.close()

tab_book = load_workbook(tables+'tabulation.xlsx')


for sheetname in tab_book.sheetnames:
    sheet=tab_book[sheetname] 
    
    sheet['A1']="Variable"
    
    sheet.column_dimensions['A'].width=60
    sheet.column_dimensions['B'].width=11
    sheet.column_dimensions['C'].width=12
    
    for r in sheet.rows:
        cells= r[:1][0]
        cells.alignment=align  
    
    for column_cells in sheet.columns:
        length = max(map(lambda cell: len(str(cell.value)) if cell.value else 0, column_cells))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width=length
        
    for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.fill=pattern       

tab_book.save(tables+'tabulation.xlsx')
############## END: TABULATION  FOR CATEGORICAL VARIABLES#################


#average_over_cat=pd.pivot_table(df, index=['region', 'year'], columns=['education', 'contact_gender'],  values=['num_chicks_sold_total', 'turn_over_total'], aggfunc=[np.mean])
#average_over_cat.to_excel(tables+'average_over_all_cat_vars.xlsx')
